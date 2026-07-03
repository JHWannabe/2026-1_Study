"""
벤더 AI 리포트(ClariSarco, D:\\영상제공\\강남\\강남_결과)의 "L3 Level (Slice #NN)" 텍스트를
OCR로 읽어 L3의 실제 위치(pubis 0% ~ liver 100% 상대위치)를 구하는 파이프라인.

TotalSegmentator(자체 세그멘테이션)보다 우선순위가 높은 소스 - 벤더 공식 판독값이며,
patient 1252461에서 TotalSegmentator 결과(65.8%)와 0.4%p 이내로 일치함을 확인함.

단계:
  1. 강남_z_bounds.xlsx 에서 patient의 series_description, z_pubis, z_liver_upper 조회
  2. 해당 시리즈 폴더에서 SeriesInstanceUID를 확인 (DICOM 헤더)
  3. 강남_DLO_Results_summary.xlsx 의 SRC_Report 하이퍼링크 중 그 UID가 포함된 리포트 이미지 경로 탐색
  4. 이미지를 OCR(Tesseract)해서 "L3 Level (Slice #NN)" 파싱
  5. 해당 Instance Number의 실제 z좌표(DICOM ImagePositionPatient)를 찾아 %로 환산

강남_z_bounds.xlsx에 있는 전체 환자에 대해 위 파이프라인을 batch 실행하고
l3_ocr_results_full.csv로 저장한다.
"""

import glob
import os
import re

import openpyxl
import pandas as pd
import pydicom
import pytesseract
from PIL import Image

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, "..", "excel")

DICOM_ROOT = r"D:\영상제공\강남\강남_axial"
RESULTS_ROOT = r"D:\영상제공\강남\강남_결과"
DLO_SUMMARY_PATH = os.path.join(RESULTS_ROOT, "강남_DLO_Results_summary.xlsx")
Z_BOUNDS_PATH = os.path.join(EXCEL_DIR, "강남_z_bounds.xlsx")
AEC_XLSX = os.path.join(EXCEL_DIR, "강남_aec_128.xlsx")
RESULT_CSV = os.path.join(EXCEL_DIR, "l3_ocr_results_full.csv")
SAVE_EVERY = 20

L3_SLICE_RE = re.compile(r"L3 Level\s*\(Slice\s*#(\d+)\)")
HYPERLINK_PATH_RE = re.compile(r'HYPERLINK\("([^"]+)"')


def idx_to_pct(idx) -> float:
    return (idx - 1) / 127 * 100


def get_series_instance_uid(patient_id, series_description: str) -> str:
    pattern = rf"{DICOM_ROOT}\{patient_id}\{series_description}\*.dcm"
    f = glob.glob(pattern)[0]
    ds = pydicom.dcmread(f, stop_before_pixels=True)
    return str(ds.SeriesInstanceUID)


def find_report_image_path(dlo_ws, patient_id, series_uid: str):
    """DLO_Results_summary의 SRC_Report 컬럼에서 series_uid가 포함된 리포트 이미지 경로 탐색"""
    headers = [c.value for c in dlo_ws[1]]
    pid_col = headers.index("PatientID") + 1
    src_col = headers.index("SRC_Report") + 1

    for row in range(2, dlo_ws.max_row + 1):
        pid = dlo_ws.cell(row=row, column=pid_col).value
        try:
            if int(pid) != int(patient_id):
                continue
        except (TypeError, ValueError):
            continue
        formula = dlo_ws.cell(row=row, column=src_col).value
        if not formula or series_uid not in str(formula):
            continue
        m = HYPERLINK_PATH_RE.search(str(formula))
        if m:
            rel_path = m.group(1).replace("./", "").replace("\\", "/")
            return os.path.join(RESULTS_ROOT, rel_path)
    return None


def ocr_l3_slice(image_path: str) -> int:
    img = Image.open(image_path)
    text = pytesseract.image_to_string(img)
    m = L3_SLICE_RE.search(text)
    if not m:
        raise ValueError(f"L3 slice not found in OCR text for {image_path}")
    return int(m.group(1))


def slice_to_pct(patient_id, series_description: str, slice_num: int, inst_min: int,
                  z_pubis: float, z_liver: float) -> float:
    """벤더 리포트의 'Slice #NN'은 raw InstanceNumber가 아니라, 그 시리즈 안에서의
    1-based 순번이다. 시리즈의 InstanceNumber가 항상 1부터 시작하는 게 아니므로
    (예: inst_min=69인 시리즈도 있음), 실제 raw InstanceNumber = inst_min + slice_num - 1
    로 변환해야 한다. (강남_z_bounds.xlsx의 inst_min 컬럼 사용)"""
    actual_instance = inst_min + slice_num - 1
    pattern = rf"{DICOM_ROOT}\{patient_id}\{series_description}\*.dcm"
    files = glob.glob(pattern)
    for f in files:
        ds = pydicom.dcmread(f, stop_before_pixels=True)
        if int(ds.InstanceNumber) == actual_instance:
            z = float(ds.ImagePositionPatient[2])
            return idx_to_pct((z - z_pubis) / (z_liver - z_pubis) * 127 + 1)
    raise ValueError(f"Instance {actual_instance} (slice #{slice_num}, inst_min={inst_min}) "
                      f"not found for patient {patient_id}")


def get_l3_pct(patient_id, zb_df: pd.DataFrame, dlo_ws) -> dict:
    zb_row = zb_df[zb_df["PatientID"] == int(patient_id)].iloc[0]
    series_description = zb_row["series_description"]
    z_pubis, z_liver = zb_row["z_pubis"], zb_row["z_liver_upper"]
    inst_min = int(zb_row["inst_min"])

    series_uid = get_series_instance_uid(patient_id, series_description)
    image_path = find_report_image_path(dlo_ws, patient_id, series_uid)
    if image_path is None or not os.path.exists(image_path):
        return {"patient_id": patient_id, "error": f"report image not found (uid={series_uid})"}

    slice_num = ocr_l3_slice(image_path)
    pct = slice_to_pct(patient_id, series_description, slice_num, inst_min, z_pubis, z_liver)
    return {"patient_id": patient_id, "series": series_description, "l3_slice": slice_num,
            "l3_pct": pct, "image_path": image_path, "error": None}


def load_zb_and_dlo():
    """강남_z_bounds.xlsx + 강남_DLO_Results_summary.xlsx 로드 (get_l3_pct 호출 전 공통 준비 단계)"""
    zb_df = pd.read_excel(Z_BOUNDS_PATH)
    dlo_wb = openpyxl.load_workbook(DLO_SUMMARY_PATH, data_only=False)
    dlo_ws = dlo_wb[dlo_wb.sheetnames[0]]
    return zb_df, dlo_ws


def select_all_samples(zb_df: pd.DataFrame):
    zb_ok_ids = set(zb_df["PatientID"])
    xls = pd.ExcelFile(AEC_XLSX)
    samples = []
    for sheet in ["M_normal", "M_low_SMI", "F_normal", "F_low_SMI"]:
        df = pd.read_excel(xls, sheet)
        for pid in df["PatientID"]:
            if pid in zb_ok_ids:
                samples.append((sheet, pid))
    return samples


def main():
    zb_df, dlo_ws = load_zb_and_dlo()
    samples = select_all_samples(zb_df)
    print(f"총 {len(samples)}명 처리 예정", flush=True)

    results = []
    for i, (sheet, pid) in enumerate(samples):
        try:
            res = get_l3_pct(pid, zb_df, dlo_ws)
        except Exception as e:
            res = {"patient_id": pid, "error": str(e)}
        res["sheet"] = sheet
        results.append(res)

        if (i + 1) % SAVE_EVERY == 0 or (i + 1) == len(samples):
            pd.DataFrame(results).to_csv(RESULT_CSV, index=False)
            n_ok = sum(1 for r in results if not r.get("error"))
            print(f"[{i+1}/{len(samples)}] 저장됨 (성공 {n_ok}/{i+1})", flush=True)

    df = pd.DataFrame(results)
    print(f"\n최종 성공: {(df['error'].isna()).sum()}/{len(df)}")
    print(f"저장: {RESULT_CSV}")


if __name__ == "__main__":
    main()
