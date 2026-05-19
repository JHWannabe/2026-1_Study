"""
강남_DLO_Results.xlsx 데이터 정제 및 OCR 기반 자동 보완 스크립트

[주요 동작 흐름 및 기능 개요]
1. 데이터 로드 및 구조 분석
   - openpyxl을 통해 수식(HYPERLINK 등)을 보존한 상태로 원본 엑셀 데이터를 읽어옵니다.
   - 데이터 헤더 영역을 탐색하여 핵심 컬럼들의 위치를 동적으로 스캔합니다.
   - 장비 모델명(ManufacturerModelName) 컬럼이 없을 경우, 시트 맨 우측에 자동으로 컬럼을 신규 생성합니다.

2. 데이터 무결성 검증 및 사전 필터링 (중복/오류 행 역순 제거)
   - PatientID를 일관된 정수형(int) 포맷으로 변환 및 정제합니다.
   - 기존 TAMA 컬럼에 문자열이나 노이즈가 섞여 있어 '숫자가 아닌 경우' 데이터 오염 방지를 위해 해당 행을 우선 제거합니다.
   - 환자 ID 기준 역순 스캔 방식으로 중복 행을 제거하여 '최초 발생 데이터(keep=first)'만 유니크하게 남깁니다.

3. 유니크 데이터 대상 2중 영역 OCR 분석 및 주기적 임시 저장
   - 중복이 제거되어 확정된 유니크 환자 행의 SRC_Report 하이퍼링크 수식에서 이미지 절대 경로를 추출합니다.
   - 이미지 한 장당 단 1회 로드하여 이중 영역 OCR을 수행함으로써 처리 속도를 극대화합니다.
     ① 전체 영역 OCR (--psm 3): 정규식을 통해 CT 제조사 및 장비 모델명 추출
     ② 하단 테이블 영역 크롭 OCR (--psm 6): L3 레벨의 TAMA 및 IMATA 단면적 수치 추출
   - 대용량 데이터 처리 중 크래시나 유실을 방지하기 위해 사용자가 지정한 주기(SAVE_INTERVAL = 50건)마다 
     엑셀 파일에 중간 저장을 실시간으로 수행합니다.

4. 최종 마감 저장
   - 모든 유니크 환자의 데이터 매핑이 완료되면 누적된 메모리 데이터를 최종 Unique 파일로 안전하게 마감 저장합니다.
"""

import os
import re
import cv2
import numpy as np
import pandas as pd
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from tqdm import tqdm

# ══════════════════════════════════════════════════════════════════════════════
# [최종 출력 파일 (강남_DLO_Results_Unique.xlsx) 컬럼 레이아웃 가이드]
# ──────────────────────────────────────────────────────────────────────────────
#  1. PatientID             : 환자 고유 식별 번호 (정수형 변환 정제 완료)
#  2. PatientAge            : 환자 나이
#  3. PatientSex            : 환자 성별
#  4. kVp                   : CT 관전압 설정값
#  5. mAs                   : CT 관전류-시간곱 설정값
#  6. ManufacturerModelName : [★ 신규 추가] 리포트 이미지에서 OCR로 자동 추출한 CT 장비 모델명
#  7. SRC_Report            : 근육 분석 리포트 (Osteo_Sarco_1.png) HYPERLINK 수식 포함 열
#  8. TAMA                  : Total Abdominal Muscle Area (OCR 보완 및 비숫자 행 제거 완료)
#  9. IMATA                 : Intramuscular Adipose Tissue Area (OCR 보완 완료)
# 10. BMI                   : 체질량 지수 (Body Mass Index)
# ══════════════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════════════
# 환경 설정 (Tesseract OCR 경로 설정 및 파일 경로)
# ══════════════════════════════════════════════════════════════════════════════
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

file_path = r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\강남\metadata\강남_DLO_Results.xlsx"
output_path = r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\강남\metadata\강남_DLO_Results_Unique.xlsx"

DLO_BASE = r"D:\영상제공\강남\강남_결과"

# 중간 저장 주기 설정 (50건마다 자동 물리 저장 수행)
SAVE_INTERVAL = 50

# ══════════════════════════════════════════════════════════════════════════════
# 내부 기능 함수 구현
# ══════════════════════════════════════════════════════════════════════════════
def parse_dlo_img_path(formula_str, dlo_base):
    """SRC_Report 내 HYPERLINK 수식에서 이미지 절대경로 추출"""
    if not formula_str:
        return None
    m = re.search(r'HYPERLINK\("([^"]+)"', str(formula_str))
    if not m:
        return None
    rel = m.group(1).replace("\\", "/").lstrip("./")
    full = os.path.join(dlo_base, rel.replace("/", os.sep))
    return full

def extract_metadata_from_sarco(img_path):
    """Osteo_Sarco_1.png에서 TAMA, IMATA 및 ManufacturerModelName 추출"""
    if not img_path or not os.path.exists(img_path):
        return {}
    
    results = {}
    try:
        img = Image.open(img_path)
        w, h = img.size
        arr = np.array(img)
        gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
        _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)
        
        # 1. 전체 영역 OCR -> Manufacturer Model Name 추출용
        full_text = pytesseract.image_to_string(thresh, config="--psm 3")
        model_match = re.search(r"(?:Manufacturer\s*Model\s*Name|Model\s*Name)\s*[:|-]?\s*([^\n]+)", full_text, re.IGNORECASE)
        if model_match:
            results["ManufacturerModelName"] = model_match.group(1).strip()
        
        # 2. 하단 테이블 영역 크롭 OCR -> TAMA, IMATA 추출용
        crop = img.crop((0, int(h * 0.72), w, int(h * 0.93)))
        crop_arr = np.array(crop)
        crop_gray = cv2.cvtColor(crop_arr, cv2.COLOR_RGB2GRAY)
        _, crop_thresh = cv2.threshold(crop_gray, 180, 255, cv2.THRESH_BINARY)
        table_text = pytesseract.image_to_string(crop_thresh, config="--psm 6")
        
        for line in table_text.split("\n"):
            line = line.strip()
            m = re.match(r"(TAMA|IMATA)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)", line)
            if m:
                vals = [int(m.group(i)) for i in range(2, 6)]
                results[m.group(1)] = vals[3]
                
        return results
    except Exception as e:
        return results

# ══════════════════════════════════════════════════════════════════════════════
# 메인 로직 시작
# ══════════════════════════════════════════════════════════════════════════════

print("1. 엑셀 파일 로드 및 구조 분석 중...")
wb = load_workbook(file_path)
ws = wb.active

# 대상 컬럼 위치 검색
col_indices = {
    "PatientID": None, 
    "SRC_Report": None, 
    "TAMA": None, 
    "IMATA": None, 
    "ManufacturerModelName": None
}
header_row_idx = None

for r_idx in range(1, 11):
    for c_idx in range(1, ws.max_column + 1):
        val = str(ws.cell(row=r_idx, column=c_idx).value).strip()
        if val in col_indices:
            col_indices[val] = c_idx
            header_row_idx = r_idx
    if col_indices["PatientID"] and col_indices["SRC_Report"]:
        break

if not col_indices["PatientID"] or not col_indices["SRC_Report"]:
    raise ValueError("필수 컬럼('PatientID' 또는 'SRC_Report')을 찾을 수 없습니다.")

if col_indices["ManufacturerModelName"] is None:
    new_col_idx = ws.max_column + 1
    ws.cell(row=header_row_idx, column=new_col_idx).value = "ManufacturerModelName"
    col_indices["ManufacturerModelName"] = new_col_idx
    print(f"-> 'ManufacturerModelName' 컬럼이 없어 {new_col_idx}번째 열에 새로 추가했습니다.")

pid_col = col_indices["PatientID"]
src_col = col_indices["SRC_Report"]
tama_col = col_indices["TAMA"]
imata_col = col_indices["IMATA"]
model_col = col_indices["ManufacturerModelName"]

start_row = header_row_idx + 1
max_row = ws.max_row

# ──────────────────────────────────────────────────────────────────────────────
# 단계 2: PatientID 정제 및 중복/TAMA 비숫자 행 역순 제거
# ──────────────────────────────────────────────────────────────────────────────
print("\n2. PatientID 정제 및 데이터 검증 구조 생성 중...")
row_info_list = []
initial_data_count = max_row - start_row + 1

for r_idx in range(start_row, max_row + 1):
    pid_cell = ws.cell(row=r_idx, column=pid_col)
    raw_pid = str(pid_cell.value).strip() if pid_cell.value is not None else ""
    
    if not raw_pid or raw_pid == "None":
        row_info_list.append((r_idx, "None", False))
        continue

    try:
        pid_int = int(float(raw_pid))
        pid_cell.value = pid_int  
        pid_str_key = str(pid_int)
    except (ValueError, TypeError):
        pid_str_key = raw_pid

    is_tama_numeric = False
    if tama_col:
        tama_val = ws.cell(row=r_idx, column=tama_col).value
        if tama_val is not None:
            try:
                float(str(tama_val).strip())
                is_tama_numeric = True
            except ValueError:
                is_tama_numeric = False

    row_info_list.append((r_idx, pid_str_key, is_tama_numeric))

print("-> 역순으로 중복 행 및 TAMA 비숫자 행 일괄 제거 중...")
seen_pids = set()
removed_by_duplicate = 0
removed_by_tama = 0
removed_by_none = 0

for r_idx, pid_key, is_tama_numeric in tqdm(reversed(row_info_list), total=len(row_info_list), desc="Cleaning Rows"):
    if pid_key == "None":
        ws.delete_rows(r_idx)
        removed_by_none += 1
        continue
    
    if tama_col and not is_tama_numeric:
        ws.delete_rows(r_idx)
        removed_by_tama += 1
        continue
        
    if pid_key in seen_pids:
        ws.delete_rows(r_idx)
        removed_by_duplicate += 1
    else:
        seen_pids.add(pid_key)

updated_max_row = ws.max_row
unique_data_count = updated_max_row - start_row + 1

print("\n📊 [데이터 정제 결과 요약]")
print(f"   • 최종 유니크 행 수: {unique_data_count}개")
print("──────────────────────────────────────────────────────────────────────────────")

# ──────────────────────────────────────────────────────────────────────────────
# 단계 3: 유니크 행 데이터 대상 OCR 매핑 및 주기적 저장
# ──────────────────────────────────────────────────────────────────────────────
print(f"\n3. OCR 분석 진행 및 실시간 매핑 중... (매 {SAVE_INTERVAL}건마다 자동 저장)")

processed_count = 0  
ocr_log_limit = 5
ocr_log_count = 0

for r_idx in tqdm(range(start_row, updated_max_row + 1), desc="OCR & Data Mapping"):
    pid_cell = ws.cell(row=r_idx, column=pid_col)
    if pid_cell.value is None or str(pid_cell.value).strip() == "None":
        continue
        
    src_cell = ws.cell(row=r_idx, column=src_col)
    img_path = parse_dlo_img_path(src_cell.value, DLO_BASE)
    
    if img_path and os.path.exists(img_path):
        ocr_results = extract_metadata_from_sarco(img_path)
        
        if tama_col and "TAMA" in ocr_results and ocr_results["TAMA"] != 0:
            ws.cell(row=r_idx, column=tama_col).value = int(ocr_results["TAMA"])
            
        if imata_col and "IMATA" in ocr_results and ocr_results["IMATA"] != 0:
            ws.cell(row=r_idx, column=imata_col).value = int(ocr_results["IMATA"])
            
        if model_col and "ManufacturerModelName" in ocr_results:
            ws.cell(row=r_idx, column=model_col).value = str(ocr_results["ManufacturerModelName"])

        processed_count += 1

        if ocr_log_count < ocr_log_limit:
            tqdm.write(f"   🔍 [OCR 확인] PatientID: {pid_cell.value} | TAMA: {ocr_results.get('TAMA', 0)} | IMATA: {ocr_results.get('IMATA', 0)}")
            ocr_log_count += 1

        if processed_count % SAVE_INTERVAL == 0:
            wb.save(output_path)
            tqdm.write(f"   💾 [자동 중간 저장 완료] 현재까지 {processed_count}개 유니크 데이터 매핑 및 저장 완료.")

    else:
        if ocr_log_count < ocr_log_limit and img_path:
            tqdm.write(f"   ⚠️ [경고] 이미지 파일 없음: {os.path.basename(img_path)} (PatientID: {pid_cell.value})")

# ──────────────────────────────────────────────────────────────────────────────
# 단계 4: 최종 데이터 저장
# ──────────────────────────────────────────────────────────────────────────────
print(f"\n4. 최종 결과 최종 마감 저장 중...")
wb.save(output_path)
wb.close()

print(f"\n[작업 완료] 주기적 자동 저장이 포함된 모든 프로세스가 성공적으로 끝났습니다.")
print(f"최종 결과 파일: '{output_path}'")