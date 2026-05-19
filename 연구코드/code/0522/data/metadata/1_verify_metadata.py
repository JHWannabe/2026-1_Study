"""
강남_DLO_Results_Unique.xlsx 데이터 기반 VB 코호트 지표 매핑 및 SMI 산출 스크립트

[주요 동작 흐름]
1. VBcohort '신장체중(예진)_전체' 시트에서 강남 환자의 신장/체중을 로드
   - 지역병원코드 기준 강남 필터링 후 환자별 전체 기록을 최신순으로 보관
2. 유니크 엑셀에 Height, Weight, BMI, SMI 컬럼 추가 후 PatientID 기준으로 채움
   - 코호트 기록 중 신장·체중이 모두 non-null인 최신 기록 사용 (범위 체크 없음)
3. BMI·SMI 결측 보완, IMATA/Height/Weight 결측 행 제거
4. 이상치 행 일괄 삭제 (Height·Weight·BMI·SMI 기준)
5. 저장
"""

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

# ══════════════════════════════════════════════════════════════════════════════
# 설정
# ══════════════════════════════════════════════════════════════════════════════
SITE_CODE = "20"  # radiation_data 강남병원 지역병원코드

# 이상치 판정 기준 (채우기 완료 후 범위 벗어난 행 삭제에 사용)
VALID_HEIGHT = (130.0, 210.0)   # cm
VALID_WEIGHT = (25.0,  200.0)   # kg
VALID_BMI    = (10.0,  55.0)
VALID_SMI    = (10.0,  80.0)    # cm²/m²

def in_range(value, bounds: tuple) -> bool:
    try:
        return bounds[0] <= float(value) <= bounds[1]
    except (TypeError, ValueError):
        return False

unique_file_path  = r"연구코드\data\강남\metadata\강남_DLO_Results_Unique.xlsx"
vb_cohort_path    = r"연구코드\data\radiation_data_260421_VBcohort.xlsx"
final_output_path = r"연구코드\data\강남\metadata\강남_DLO_Results_SMI.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# 1. VB 코호트 로드
# ══════════════════════════════════════════════════════════════════════════════
print("1. VB 코호트 파일 로드 및 검색 인덱스 생성 중...")
try:
    df_cohort = pd.read_excel(vb_cohort_path, sheet_name='신장체중(예진)_전체', dtype=str)
    df_cohort = df_cohort[df_cohort['지역병원코드'] == SITE_CODE].copy()
    df_cohort['신장'] = pd.to_numeric(df_cohort['신장'], errors='coerce')
    df_cohort['체중'] = pd.to_numeric(df_cohort['체중'], errors='coerce')
    df_cohort['PatientID_key'] = df_cohort['연구등록번호'].str.strip().str.split('.').str[0]
    df_cohort = df_cohort.sort_values('내원연월', ascending=False)
    cohort_dict = {  # type: ignore[var-annotated]
        pid: records.to_dict('records')
        for pid, records in df_cohort.groupby('PatientID_key', sort=False)
    }
    print(f"-> 강남 예진 데이터 {len(cohort_dict)}명 확보.")
except Exception as e:
    raise IOError(f"VBcohort.xlsx 파일을 읽는 중 오류가 발생했습니다: {e}")

# ══════════════════════════════════════════════════════════════════════════════
# 2. 유니크 엑셀 로드 및 컬럼 구조 파악
# ══════════════════════════════════════════════════════════════════════════════
print("\n2. 원본 유니크 엑셀 파일 로드 및 구조 분석 중...")
wb = load_workbook(unique_file_path)
ws = wb.active
assert ws is not None, "활성 시트를 찾을 수 없습니다."

col_indices = {
    "PatientID": None,
    "TAMA":      None,
    "IMATA":     None,
    "BMI":       None,
    "Height":    None,
    "Weight":    None,
    "SMI":       None,
}
header_row_idx = None

max_col: int = ws.max_column or 1  # type: ignore[assignment]
for r_idx in range(1, 11):
    for c_idx in range(1, max_col + 1):
        val = str(ws.cell(row=r_idx, column=c_idx).value).strip()
        if val in col_indices:
            col_indices[val] = c_idx
            header_row_idx = r_idx
    if col_indices["PatientID"] and col_indices["TAMA"]:
        break

if not col_indices["PatientID"] or header_row_idx is None:
    raise ValueError("시트 안에서 'PatientID' 컬럼을 찾을 수 없습니다.")

new_columns = ["Height", "Weight", "SMI"]
if col_indices["BMI"] is None:
    new_columns.insert(0, "BMI")

current_max_col: int = ws.max_column or max_col  # type: ignore[assignment]
for col_name in new_columns:
    if col_indices[col_name] is None:
        current_max_col += 1
        ws.cell(row=header_row_idx, column=current_max_col).value = col_name  # type: ignore[union-attr]
        col_indices[col_name] = current_max_col
        print(f"-> '{col_name}' 컬럼 추가 ({current_max_col}번째 열)")

pid_col    = col_indices["PatientID"]
tama_col   = col_indices["TAMA"]
imata_col  = col_indices["IMATA"]
bmi_col    = col_indices["BMI"]
height_col = col_indices["Height"]
weight_col = col_indices["Weight"]
smi_col    = col_indices["SMI"]

assert pid_col is not None and tama_col is not None and imata_col is not None

# ══════════════════════════════════════════════════════════════════════════════
# 3. PatientID 매칭, 신장/체중/BMI/SMI 기입 (범위 체크 없이 전부 채움)
# ══════════════════════════════════════════════════════════════════════════════
print("\n3. VB 코호트 매칭 및 지표 기입 중...")

for r_idx in tqdm(range(header_row_idx + 1, ws.max_row + 1), desc="Mapping & Calculating"):
    pid_val = ws.cell(row=r_idx, column=pid_col).value
    if pid_val is None or str(pid_val).strip() == "None":
        continue

    pid_key = str(pid_val).strip().split('.')[0]

    raw_tama  = ws.cell(row=r_idx, column=tama_col).value
    raw_imata = ws.cell(row=r_idx, column=imata_col).value
    tama  = float(raw_tama)  if raw_tama  is not None else 0.0  # type: ignore[arg-type]
    imata = float(raw_imata) if raw_imata is not None else 0.0  # type: ignore[arg-type]

    if pid_key not in cohort_dict:
        continue

    # 최신순 기록 중 신장·체중이 non-null이고 이상치 범위 내인 첫 번째 기록 사용
    height, weight = None, None
    for record in cohort_dict[pid_key]:
        h = record.get('신장(cm)') or record.get('신장')
        w = record.get('체중(kg)') or record.get('체중')
        if h is None or pd.isna(h) or w is None or pd.isna(w):
            continue
        if not in_range(h, VALID_HEIGHT) or not in_range(w, VALID_WEIGHT):
            continue
        h_m = float(h) / 100.0 if float(h) > 3.0 else float(h)
        if h_m <= 0 or not in_range(float(w) / (h_m ** 2), VALID_BMI):
            continue
        height, weight = h, w
        break

    if height is not None and height_col is not None:
        ws.cell(row=r_idx, column=height_col).value = float(height)  # type: ignore[union-attr]
    if weight is not None and weight_col is not None:
        ws.cell(row=r_idx, column=weight_col).value = float(weight)  # type: ignore[union-attr]

    if height is not None and weight is not None and bmi_col is not None:
        height_m = float(height) / 100.0 if float(height) > 3.0 else float(height)
        if height_m > 0:
            ws.cell(row=r_idx, column=bmi_col).value = round(float(weight) / (height_m ** 2), 2)  # type: ignore[union-attr]

    if height is not None and smi_col is not None:
        try:
            height_m = float(height) / 100.0 if float(height) > 3.0 else float(height)
            if height_m > 0:
                ws.cell(row=r_idx, column=smi_col).value = round((tama - imata) / (height_m ** 2), 2)  # type: ignore[union-attr]
        except ZeroDivisionError:
            pass

# ══════════════════════════════════════════════════════════════════════════════
# 4. BMI 결측 보완 (Height·Weight 있고 BMI 없는 행)
# ══════════════════════════════════════════════════════════════════════════════
print("\n4. BMI 결측치 보완 중...")
bmi_filled = 0
for r_idx in range(header_row_idx + 1, ws.max_row + 1):  # type: ignore[operator]
    if bmi_col is None or height_col is None or weight_col is None:
        break
    if ws.cell(row=r_idx, column=bmi_col).value is not None:
        continue
    height_val = ws.cell(row=r_idx, column=height_col).value
    weight_val = ws.cell(row=r_idx, column=weight_col).value
    if height_val is None or weight_val is None:
        continue
    try:
        h = float(height_val)  # type: ignore[arg-type]
        w = float(weight_val)  # type: ignore[arg-type]
        height_m = h / 100.0 if h > 3.0 else h
        if height_m > 0:
            ws.cell(row=r_idx, column=bmi_col).value = round(w / (height_m ** 2), 2)  # type: ignore[union-attr]
            bmi_filled += 1
    except (ValueError, ZeroDivisionError):
        pass
print(f"-> BMI 보완 완료: {bmi_filled}명")

# ══════════════════════════════════════════════════════════════════════════════
# 5. IMATA 결측 행 제거
# ══════════════════════════════════════════════════════════════════════════════
print("\n5. IMATA 결측 행 제거 중...")
removed = 0
for r_idx in range(ws.max_row, header_row_idx, -1):  # type: ignore[operator]
    val = ws.cell(row=r_idx, column=imata_col).value
    if val is None or str(val).strip() in ("", "None"):
        ws.delete_rows(r_idx)
        removed += 1
print(f"-> {removed}행 제거 완료")

# ══════════════════════════════════════════════════════════════════════════════
# 6. Height 또는 Weight 결측 행 제거
# ══════════════════════════════════════════════════════════════════════════════
print("\n6. Height·Weight 결측 행 제거 중...")
removed_hw = 0
for r_idx in range(ws.max_row, header_row_idx, -1):  # type: ignore[operator]
    h_val = ws.cell(row=r_idx, column=height_col).value if height_col else None
    w_val = ws.cell(row=r_idx, column=weight_col).value if weight_col else None
    if h_val is None or w_val is None:
        ws.delete_rows(r_idx)
        removed_hw += 1
print(f"-> {removed_hw}행 제거 완료")

# ══════════════════════════════════════════════════════════════════════════════
# 7. SMI 결측 보완 (Height·TAMA·IMATA 있고 SMI 없는 행)
# ══════════════════════════════════════════════════════════════════════════════
print("\n7. SMI 결측치 보완 중...")
smi_filled = 0
for r_idx in range(header_row_idx + 1, ws.max_row + 1):  # type: ignore[operator]
    if smi_col is None or height_col is None:
        break
    if ws.cell(row=r_idx, column=smi_col).value is not None:
        continue
    height_val = ws.cell(row=r_idx, column=height_col).value
    tama_val   = ws.cell(row=r_idx, column=tama_col).value
    imata_val  = ws.cell(row=r_idx, column=imata_col).value
    if height_val is None or tama_val is None or imata_val is None:
        continue
    try:
        h = float(height_val)  # type: ignore[arg-type]
        height_m = h / 100.0 if h > 3.0 else h
        if height_m > 0:
            ws.cell(row=r_idx, column=smi_col).value = round(  # type: ignore[union-attr]
                (float(tama_val) - float(imata_val)) / (height_m ** 2), 2  # type: ignore[arg-type]
            )
            smi_filled += 1
    except (ValueError, ZeroDivisionError):
        pass
print(f"-> SMI 보완 완료: {smi_filled}명")

# ══════════════════════════════════════════════════════════════════════════════
# 8. 이상치 행 삭제 (Height·Weight·BMI·SMI 기준, 아래부터 순회)
# ══════════════════════════════════════════════════════════════════════════════
print("\n8. 이상치 행 삭제 중...")
removed_outlier = 0
for r_idx in range(ws.max_row, header_row_idx, -1):  # type: ignore[operator]
    h_val   = ws.cell(row=r_idx, column=height_col).value if height_col else None
    w_val   = ws.cell(row=r_idx, column=weight_col).value if weight_col else None
    bmi_val = ws.cell(row=r_idx, column=bmi_col).value   if bmi_col    else None
    smi_val = ws.cell(row=r_idx, column=smi_col).value   if smi_col    else None

    is_outlier = (
        (h_val   is not None and not in_range(h_val,   VALID_HEIGHT)) or
        (w_val   is not None and not in_range(w_val,   VALID_WEIGHT)) or
        (bmi_val is not None and not in_range(bmi_val, VALID_BMI))    or
        (smi_val is not None and not in_range(smi_val, VALID_SMI))
    )
    if is_outlier:
        ws.delete_rows(r_idx)
        removed_outlier += 1
print(f"-> 이상치 {removed_outlier}행 삭제 완료")

# ══════════════════════════════════════════════════════════════════════════════
# 9. 저장
# ══════════════════════════════════════════════════════════════════════════════
print(f"\n9. 최종 파일 저장 중...")
wb.save(final_output_path)
wb.close()

print(f"\n[완료] SMI 산출 결과 저장: '{final_output_path}'")
