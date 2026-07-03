"""
[전체 목적]
PHASE 0: DLO 결과 엑셀 중복 제거 + OCR 메타데이터 보완
  → {SITE}_DLO_Results.xlsx → {SITE}_DLO_Results_Unique.xlsx
"""

import os
import re
import cv2
import numpy as np
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from tqdm import tqdm

# ── 공통 환경 설정 ────────────────────────────────────────────────────────────
# 여기만 바꾸면 전체 경로·코드가 자동으로 연동된다.

SITE = "강남"   # "강남" 또는 "신촌"

BASE_DIR  = rf"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\{SITE}\metadata"
DLO_BASE  = rf"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\{SITE}\metadata"
file_path   = os.path.join(BASE_DIR, f"{SITE}_DLO_Results.xlsx")
output_path = os.path.join(BASE_DIR, f"{SITE}_DLO_Results_Unique.xlsx")

SAVE_INTERVAL = 50

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
_SERIES_PAT = re.compile(r"Series\s*Desc[^\n]*?:\s*([^:\n]+)", re.IGNORECASE)
_BOLD_CENTER = Font(bold=True), Alignment(horizontal="center")


# ── 공통 함수 ─────────────────────────────────────────────────────────────────

def parse_dlo_img_path(formula_str, dlo_base):
    if not formula_str:
        return None
    m = re.search(r'HYPERLINK\("([^"]+)"', str(formula_str))
    if not m:
        return None
    rel  = m.group(1).replace("\\", "/").lstrip("./")
    full = os.path.join(dlo_base, rel.replace("/", os.sep))
    return full


def _remove_horizontal_lines(binary_img: np.ndarray) -> np.ndarray:
    """표 구분선 등 수평선을 제거하여 겹쳐진 글자(+, - 등)의 OCR 인식률을 높임"""
    _, w = binary_img.shape
    inv = cv2.bitwise_not(binary_img)
    # 이미지 너비의 1/5 이상 되는 수평 획만 선으로 간주
    kernel_w = max(w // 5, 50)
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_w, 1))
    h_lines  = cv2.morphologyEx(inv, cv2.MORPH_OPEN, h_kernel, iterations=2)
    return cv2.bitwise_not(cv2.subtract(inv, h_lines))


def _detect_plus_after_as(thresh_clean: np.ndarray) -> bool:
    """'AS' 단어 바로 오른쪽 픽셀 영역을 검사해 '+' 세로획 잔재 여부를 반환"""
    try:
        data = pytesseract.image_to_data(
            thresh_clean, config="--psm 3",
            output_type=pytesseract.Output.DICT,
        )
        img_h, img_w = thresh_clean.shape
        for i in range(len(data["text"]) - 1, -1, -1):
            if str(data["text"][i]).strip().upper() != "AS":
                continue
            if int(data["conf"][i]) <= 0:
                continue
            x_right   = int(data["left"][i]) + int(data["width"][i])
            y_top     = max(int(data["top"][i]) - 3, 0)
            y_bot     = min(int(data["top"][i]) + int(data["height"][i]) + 3, img_h)
            x_end     = min(x_right + 35, img_w)
            if x_right >= x_end:
                continue
            region = thresh_clean[y_top:y_bot, x_right:x_end]
            if region.size == 0:
                continue
            # 어두운 픽셀(=글자) 비율이 10% 초과이면 세로획 존재로 판단
            return float(np.sum(region == 0)) / region.size > 0.10
    except Exception:
        pass
    return False


def extract_metadata_from_sarco(img_path):
    if not img_path or not os.path.exists(img_path):
        return {}
    results = {}
    try:
        img = Image.open(img_path)
        w, h = img.size
        arr  = np.array(img)
        gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
        _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)

        # 수평선 제거 후 OCR
        thresh_clean = _remove_horizontal_lines(thresh)
        full_text    = pytesseract.image_to_string(thresh_clean, config="--psm 3")

        model_match = re.search(
            r"(?:Manufacturer\s*Model\s*Name|Model\s*Name)\s*[:|-]?\s*([^\n]+)",
            full_text, re.IGNORECASE
        )
        if model_match:
            model_val = model_match.group(1).strip()

            if re.search(r"SOMATOM\s+Definition\s+AS", model_val, re.IGNORECASE):
                # 1단계: OCR가 세로획을 문자로 읽은 경우 → 직접 치환
                corrected = re.sub(
                    r"(SOMATOM\s+Definition\s+AS)\s*[|lI1/\\]+\s*$",
                    r"\1+",
                    model_val,
                    flags=re.IGNORECASE,
                )
                if "+" in corrected:
                    model_val = corrected
                # 2단계: 세로획을 완전히 무시한 경우 → 픽셀 검사
                elif re.search(r"SOMATOM\s+Definition\s+AS$", model_val, re.IGNORECASE):
                    if _detect_plus_after_as(thresh_clean):
                        model_val = model_val.rstrip() + "+"

            results["ManufacturerModelName"] = model_val

        series_match = _SERIES_PAT.search(full_text)
        if series_match:
            val = series_match.group(1).strip()
            val = re.split(r'\s*(?:CTD\S*|Slice\S*)', val, flags=re.IGNORECASE)[0].strip()
            results["SeriesDescription"] = val

        mas_match = re.search(
            r"(?:(?:Effective\s+)?mAs\s*[:\-]?\s*(\d+(?:\.\d+)?)"
            r"|(\d+(?:\.\d+)?)\s*mAs)",
            full_text, re.IGNORECASE,
        )
        if mas_match:
            try:
                results["mAs"] = float(mas_match.group(1) or mas_match.group(2))
            except (ValueError, TypeError):
                pass

        crop      = img.crop((0, int(h * 0.72), w, int(h * 0.93)))
        crop_arr  = np.array(crop)
        crop_gray = cv2.cvtColor(crop_arr, cv2.COLOR_RGB2GRAY)
        _, crop_thresh = cv2.threshold(crop_gray, 180, 255, cv2.THRESH_BINARY)
        table_text = pytesseract.image_to_string(crop_thresh, config="--psm 6")

        for line in table_text.split("\n"):
            line = line.strip()
            m = re.match(r"(TAMA|IMATA)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)", line)
            if m:
                vals = [int(m.group(i)) for i in range(2, 6)]
                results[m.group(1)] = vals[3]

        return results
    except Exception:
        return results


def drop_blank_columns(ws, header_row: int | None):
    if header_row is None:
        return
    for c_idx in range(ws.max_column, 0, -1):
        if ws.cell(row=header_row, column=c_idx).value in (None, ""):
            ws.delete_cols(c_idx)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 0 단계 함수
# ══════════════════════════════════════════════════════════════════════════════

def step0_1_load_excel(file_path: str):
    """엑셀 로드 및 컬럼 위치 탐색, 누락 컬럼 추가"""
    print("\n[0-1] 엑셀 파일 로드 및 구조 분석 중...")
    wb = load_workbook(file_path)
    ws = wb.active
    assert ws is not None, "활성 시트를 찾을 수 없습니다."

    col_indices: dict[str, int | None] = {
        "PatientID": None, "SRC_Report": None,
        "TAMA": None, "IMATA": None,
        "Manufacturer": None, "Series_Desc": None, "kVp": None, "mAs": None,
    }
    header_row_idx = None

    for r_idx in range(1, 11):
        for c_idx in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r_idx, column=c_idx).value).strip()
            if val in col_indices:
                col_indices[val] = c_idx
                header_row_idx   = r_idx
        if col_indices["PatientID"] and col_indices["SRC_Report"]:
            break

    if not col_indices["PatientID"] or not col_indices["SRC_Report"]:
        raise ValueError("필수 컬럼('PatientID' 또는 'SRC_Report')을 찾을 수 없습니다.")
    assert header_row_idx is not None

    _TARGET_COL = {"Manufacturer": 8, "Series_Desc": 9}
    for col_name in ("Manufacturer", "Series_Desc"):
        if col_indices[col_name] is None:
            target = _TARGET_COL[col_name]
            ws.insert_cols(target)
            for k in list(col_indices):
                v = col_indices[k]
                if v is not None and v >= target:
                    col_indices[k] = v + 1
            cell = ws.cell(row=header_row_idx, column=target)  # type: ignore[arg-type]
            cell.value = col_name  # type: ignore[assignment]
            cell.font, cell.alignment = _BOLD_CENTER
            col_indices[col_name] = target
            print(f"-> '{col_name}' 컬럼 없음 → {target}번째 열({'G' if target == 7 else 'H'})에 삽입")

    if col_indices["mAs"] is None:
        kvp_col = col_indices["kVp"]
        target = (kvp_col + 1) if kvp_col is not None else (ws.max_column + 1)
        if kvp_col is not None:
            ws.insert_cols(target)
            for k in list(col_indices):
                v = col_indices[k]
                if v is not None and v >= target:
                    col_indices[k] = v + 1
        cell = ws.cell(row=header_row_idx, column=target)  # type: ignore[arg-type]
        cell.value = "mAs"  # type: ignore[assignment]
        cell.font, cell.alignment = _BOLD_CENTER
        col_indices["mAs"] = target
        loc = "kVp 옆" if kvp_col is not None else "마지막 열"
        print(f"-> 'mAs' 컬럼 없음 → {target}번째 열({loc})에 삽입")

    return wb, ws, col_indices, header_row_idx


def step0_2_ocr_metadata(wb, ws, col_indices: dict, header_row_idx: int, output_path: str):
    """전체 행 OCR 메타데이터 추출 및 주기적 자동 저장"""
    print(f"\n[0-2] 전체 행 OCR 분석 중... (매 {SAVE_INTERVAL}건마다 자동 저장)")
    pid_col    = col_indices["PatientID"]
    src_col    = col_indices["SRC_Report"]
    tama_col   = col_indices["TAMA"]
    imata_col  = col_indices["IMATA"]
    model_col  = col_indices["Manufacturer"]
    series_col = col_indices["Series_Desc"]
    mas_col    = col_indices["mAs"]
    start_row  = header_row_idx + 1
    max_row    = ws.max_row

    processed = ocr_log = 0
    for r_idx in tqdm(range(start_row, max_row + 1), desc="OCR & Metadata"):
        pid_cell = ws.cell(row=r_idx, column=pid_col)
        if pid_cell.value is None or str(pid_cell.value).strip() in ("", "None"):
            continue
        img_path = parse_dlo_img_path(ws.cell(row=r_idx, column=src_col).value, DLO_BASE)
        if img_path and os.path.exists(img_path):
            ocr = extract_metadata_from_sarco(img_path)
            if tama_col   and "TAMA"  in ocr and ocr["TAMA"]  != 0:
                ws.cell(row=r_idx, column=tama_col).value   = int(ocr["TAMA"])   # type: ignore[assignment]
            if imata_col  and "IMATA" in ocr and ocr["IMATA"] != 0:
                ws.cell(row=r_idx, column=imata_col).value  = int(ocr["IMATA"])  # type: ignore[assignment]
            if model_col  and "ManufacturerModelName" in ocr:
                ws.cell(row=r_idx, column=model_col).value  = str(ocr["ManufacturerModelName"])  # type: ignore[assignment]
            if series_col and "SeriesDescription" in ocr:
                ws.cell(row=r_idx, column=series_col).value = str(ocr["SeriesDescription"])  # type: ignore[assignment]
            if mas_col and "mAs" in ocr:
                ws.cell(row=r_idx, column=mas_col).value = float(ocr["mAs"])  # type: ignore[assignment]
            processed += 1
            if ocr_log < 5:
                tqdm.write(f"   [OCR] PatientID={pid_cell.value} "
                           f"TAMA={ocr.get('TAMA', 0)} IMATA={ocr.get('IMATA', 0)} "
                           f"mAs={ocr.get('mAs', 'N/A')}")
                ocr_log += 1
            if processed % SAVE_INTERVAL == 0:
                wb.save(output_path)
                tqdm.write(f"   [자동 저장] {processed}건 완료")
        elif img_path and ocr_log < 5:
            tqdm.write(f"   [경고] 이미지 없음: {os.path.basename(img_path)} "
                       f"(PatientID={pid_cell.value})")

def step0_3_save(wb, ws, header_row_idx: int, output_path: str):
    """빈 컬럼 제거 후 최종 저장"""
    print(f"\n[0-5] 최종 저장 중...")
    drop_blank_columns(ws, header_row_idx)
    wb.save(output_path)
    wb.close()
    print(f"[PHASE 0 완료] '{output_path}'")


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 오케스트레이터
# ══════════════════════════════════════════════════════════════════════════════

def run_phase0():
    print("=" * 78)
    print(f"PHASE 0 [{SITE}]: 중복 제거 및 OCR 메타데이터 보완")
    print("=" * 78)

    wb, ws, col_indices, header_row_idx = step0_1_load_excel(file_path)
    step0_2_ocr_metadata(wb, ws, col_indices, header_row_idx, output_path)
    step0_3_save(wb, ws, header_row_idx, output_path)


if __name__ == "__main__":
    run_phase0()

    print("\n" + "=" * 78)
    print(f"[전체 파이프라인 완료] {SITE} 데이터 처리 종료")
    print("=" * 78)
