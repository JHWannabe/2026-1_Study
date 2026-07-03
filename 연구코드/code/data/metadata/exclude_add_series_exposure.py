"""
신촌_DLO_Results_SMI_kVp100.xlsx kVp_100 시트의 mAs 칼럼 채우기

SRC_Report HYPERLINK에서 이미지 경로를 얻어 OCR로 mAs 값을 추출한다.
숫자가 아니면 N/A로 저장.
"""

import os
import re
import cv2
import numpy as np
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from tqdm import tqdm

# ── 설정 ──────────────────────────────────────────────────────────────────────

SINCHON_PATH  = r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\신촌\metadata\신촌_DLO_Results_SMI_kVp100.xlsx"
SINCHON_BASE  = r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\신촌\metadata"
SINCHON_SHEET = "kVp_100"

SAVE_INTERVAL = 50

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

_EXPOSURE_PAT = re.compile(r"(\d+(?:[.,]\d+|\s+\d{1,2})?)\s*mAs", re.IGNORECASE)

# ── OCR 헬퍼 함수 ─────────────────────────────────────────────────────────────

def _parse_hyperlink_path(formula_str, base_dir) -> str | None:
    """HYPERLINK 수식에서 이미지 절대 경로를 반환."""
    if not formula_str:
        return None
    m = re.search(r'HYPERLINK\("([^"]+)"', str(formula_str))
    if not m:
        return None
    rel = m.group(1).replace("\\", "/").lstrip("./")
    return os.path.join(base_dir, rel.replace("/", os.sep))


def _remove_horizontal_lines(binary_img: np.ndarray) -> np.ndarray:
    _, w = binary_img.shape
    inv = cv2.bitwise_not(binary_img)
    kernel_w = max(w // 5, 50)
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_w, 1))
    h_lines  = cv2.morphologyEx(inv, cv2.MORPH_OPEN, h_kernel, iterations=2)
    return cv2.bitwise_not(cv2.subtract(inv, h_lines))


def extract_exposure(img_path: str) -> float | None:
    """이미지에서 mAs(Exposure) 값을 OCR로 추출. 실패하면 None."""
    if not img_path or not os.path.exists(img_path):
        return None
    try:
        img = Image.open(img_path)
        arr  = np.array(img)
        gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
        _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)
        thresh_clean = _remove_horizontal_lines(thresh)
        full_text    = pytesseract.image_to_string(thresh_clean, config="--psm 3")

        m = _EXPOSURE_PAT.search(full_text)
        if m:
            raw = re.sub(r'[,\s]+', '.', m.group(1).strip())
            return float(raw)
    except Exception:
        pass
    return None


# ── 유틸 ──────────────────────────────────────────────────────────────────────

def find_col(ws, header_row_idx: int, names: list[str]) -> int | None:
    for c in range(1, ws.max_column + 1):
        if str(ws.cell(row=header_row_idx, column=c).value or "").strip() in names:
            return c
    return None


# ── 메인 ─────────────────────────────────────────────────────────────────────

def run_sinchon():
    """신촌_DLO_Results_SMI_kVp100.xlsx kVp_100 시트의 mAs 칼럼을 OCR로 채운다."""
    print(f"[로드] {SINCHON_PATH}")
    wb = load_workbook(SINCHON_PATH)
    ws = wb[SINCHON_SHEET]

    header_row = None
    for r in range(1, 6):
        for c in range(1, ws.max_column + 1):
            if str(ws.cell(row=r, column=c).value or "").strip() == "PatientID":
                header_row = r
                break
        if header_row:
            break
    if header_row is None:
        raise ValueError("kVp_100 시트에서 PatientID 헤더를 찾을 수 없습니다.")

    mas_col = find_col(ws, header_row, ["mAs"])
    src_col = find_col(ws, header_row, ["SRC_Report"])
    if mas_col is None:
        raise ValueError("mAs 칼럼을 찾을 수 없습니다.")
    if src_col is None:
        raise ValueError("SRC_Report 칼럼을 찾을 수 없습니다.")

    print(f"-> 헤더 {header_row}행 | mAs={mas_col}열 | SRC_Report={src_col}열")
    print(f"-> 처리 대상: {ws.max_row - header_row}행")

    processed = not_found = 0
    for r in tqdm(range(header_row + 1, ws.max_row + 1), desc="OCR"):
        src_val = ws.cell(row=r, column=src_col).value
        if not src_val:
            continue

        img_path = _parse_hyperlink_path(str(src_val), SINCHON_BASE)
        if not img_path or not os.path.exists(img_path):
            not_found += 1
            ws.cell(row=r, column=mas_col).value = "N/A"  # type: ignore[assignment]
            continue

        exposure = extract_exposure(img_path)
        if isinstance(exposure, float):
            ws.cell(row=r, column=mas_col).value = exposure  # type: ignore[assignment]
        else:
            ws.cell(row=r, column=mas_col).value = "N/A"  # type: ignore[assignment]

        processed += 1
        if processed % SAVE_INTERVAL == 0:
            wb.save(SINCHON_PATH)
            tqdm.write(f"   [자동 저장] {processed}건 완료")

    wb.save(SINCHON_PATH)
    wb.close()
    print(f"\n[완료] OCR 처리: {processed}건 | 이미지 없음(N/A): {not_found}건")
    print(f"-> '{SINCHON_PATH}'")


if __name__ == "__main__":
    run_sinchon()
