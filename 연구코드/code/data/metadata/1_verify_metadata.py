"""
[전체 목적]
0_unique.py가 생성한 강남_DLO_Results_Unique.xlsx에
VB 코호트(예진 검사 기록)의 신장·체중을 매핑해
BMI와 SMI(Skeletal Muscle Index)를 계산하고 최종 정제된 파일을 저장한다.

[처리 흐름]
  STEP 1. VB 코호트 로드 → 강남 병원 필터링 → 환자별 기록 딕셔너리 생성
  STEP 2. Unique 엑셀 로드 → 컬럼 위치 탐색 → 신규 컬럼 추가
  STEP 3. PatientID 매칭 → 신장/체중/BMI/SMI 기입
  STEP 4. BMI 결측 보완
  STEP 5. IMATA 결측 행 제거
  STEP 6. Height·Weight 결측 행 제거
  STEP 7. SMI 결측 보완
  STEP 8. 이상치 행 삭제
  STEP 9. 최종 저장

[핵심 수식]
  BMI = 체중(kg) / 신장(m)²
  SMI = (TAMA - IMATA) / 신장(m)²   [단위: cm²/m²]
    - TAMA  : 복벽 근육 전체 단면적
    - IMATA : 근육 내 지방 단면적
    - TAMA - IMATA : 순수 근육 단면적(SMATA, Skeletal Muscle AT Area)
    - 신장²로 나눠 체격(키) 보정 → 체형에 무관한 상대적 근육량 지표

[용어 설명]
  VB 코호트    : 예진(건강검진) 데이터베이스. 신장·체중 등 신체계측 정보 포함.
  지역병원코드 : VB 코호트에서 각 병원을 구분하는 코드. 강남 = "20".
  최신순 기록  : 동일 환자의 여러 방문 기록 중 가장 최근 날짜 기록을 우선 사용.
  이상치(outlier) : 생리적으로 불가능하거나 입력 오류로 추정되는 값.
                    예: 신장 130cm 미만 또는 210cm 초과.
"""

import pandas as pd
from openpyxl import load_workbook
from tqdm import tqdm

# ── 설정 ─────────────────────────────────────────────────────────────────────

# 강남 병원의 VB 코호트 내 지역병원코드
SITE_CODE = "10"

# 이상치 판정 기준: 이 범위를 벗어나는 값을 가진 행을 제거한다
VALID_HEIGHT = (130.0, 210.0)   # cm
VALID_WEIGHT = (25.0,  200.0)   # kg
VALID_BMI    = (10.0,  55.0)    # kg/m²
VALID_SMI    = (10.0,  80.0)    # cm²/m² (근감소증 기준선보다 훨씬 낮거나 높은 값 제거)


def in_range(value, bounds: tuple) -> bool:
    """값이 [하한, 상한] 범위 내에 있으면 True 반환. 변환 불가 값은 False."""
    try:
        return bounds[0] <= float(value) <= bounds[1]
    except (TypeError, ValueError):
        return False


# 파일 경로 (상대경로: 스크립트 위치 기준)
unique_file_path  = r"연구코드\data\신촌\metadata\신촌_DLO_Results_Unique.xlsx"
vb_cohort_path    = r"연구코드\data\radiation_data_260421_VBcohort.xlsx"
final_output_path = r"연구코드\data\신촌\metadata\신촌_DLO_Results_SMI.xlsx"


# ── STEP 1: VB 코호트 로드 ────────────────────────────────────────────────────
# '신장체중(예진)_전체' 시트에서 강남 병원 데이터만 필터링하고,
# PatientID별로 방문 기록 전체를 딕셔너리에 담는다.
# 최신 기록을 우선 사용하기 위해 내원연월 내림차순 정렬을 먼저 수행한다.

print("1. VB 코호트 파일 로드 및 검색 인덱스 생성 중...")
try:
    # dtype=str: 환자 ID가 숫자처럼 보여도 앞자리 0이 잘리지 않도록 문자열로 읽음
    df_cohort = pd.read_excel(vb_cohort_path, sheet_name='신장체중(예진)_전체', dtype=str)
    # 강남 병원 데이터만 선택
    df_cohort = df_cohort[df_cohort['지역병원코드'] == SITE_CODE].copy()
    # 신장·체중 수치 변환 (문자열 → float, 변환 불가 값은 NaN)
    df_cohort['신장'] = pd.to_numeric(df_cohort['신장'], errors='coerce')
    df_cohort['체중'] = pd.to_numeric(df_cohort['체중'], errors='coerce')
    # 연구등록번호에 ".0" 같은 소수점이 붙어 있을 수 있어 정수 부분만 사용
    df_cohort['PatientID_key'] = df_cohort['연구등록번호'].str.strip().str.split('.').str[0]
    # 최신 기록이 앞에 오도록 내림차순 정렬
    df_cohort = df_cohort.sort_values('내원연월', ascending=False)
    # PatientID → [기록1, 기록2, ...] 딕셔너리 (최신순 정렬된 기록 리스트)
    cohort_dict = {
        pid: records.to_dict('records')
        for pid, records in df_cohort.groupby('PatientID_key', sort=False)
    }
    print(f"-> 강남 예진 데이터 {len(cohort_dict)}명 확보.")
except Exception as e:
    raise IOError(f"VBcohort.xlsx 파일을 읽는 중 오류가 발생했습니다: {e}")


# ── STEP 2: Unique 엑셀 로드 및 컬럼 구조 파악 ───────────────────────────────
# openpyxl을 사용하는 이유: 셀 단위로 직접 읽고 쓸 수 있어
# 새 컬럼 추가 및 행 삭제를 수식·서식 보존 상태에서 처리할 수 있다.

print("\n2. 원본 유니크 엑셀 파일 로드 및 구조 분석 중...")
wb = load_workbook(unique_file_path)
ws = wb.active
assert ws is not None, "활성 시트를 찾을 수 없습니다."

# 탐색할 컬럼명 목록
col_indices = {
    "PatientID": None,
    "TAMA":      None,
    "IMATA":     None,
    "BMI":       None,
    "Height":    None,
    "Weight":    None,
    "SMI":       None,
}
header_row_idx = None

max_col: int = ws.max_column or 1
# 첫 10행 안에서 헤더 행을 찾는다
for r_idx in range(1, 11):
    for c_idx in range(1, max_col + 1):
        val = str(ws.cell(row=r_idx, column=c_idx).value).strip()
        if val in col_indices:
            col_indices[val] = c_idx
            header_row_idx   = r_idx
    if col_indices["PatientID"] and col_indices["TAMA"]:
        break

if not col_indices["PatientID"] or header_row_idx is None:
    raise ValueError("시트 안에서 'PatientID' 컬럼을 찾을 수 없습니다.")

# 없는 컬럼(BMI, Height, Weight, SMI)을 시트 오른쪽에 추가
# BMI가 이미 있으면 새 컬럼 목록에서 제외
new_columns = ["Height", "Weight", "SMI"]
if col_indices["BMI"] is None:
    new_columns.insert(0, "BMI")

current_max_col: int = ws.max_column or max_col
for col_name in new_columns:
    if col_indices[col_name] is None:
        current_max_col += 1
        ws.cell(row=header_row_idx, column=current_max_col).value = col_name
        col_indices[col_name] = current_max_col
        print(f"-> '{col_name}' 컬럼 추가 ({current_max_col}번째 열)")

# 자주 쓰이는 열 인덱스를 변수에 저장
pid_col    = col_indices["PatientID"]
tama_col   = col_indices["TAMA"]
imata_col  = col_indices["IMATA"]
bmi_col    = col_indices["BMI"]
height_col = col_indices["Height"]
weight_col = col_indices["Weight"]
smi_col    = col_indices["SMI"]

assert pid_col is not None and tama_col is not None and imata_col is not None


# ── STEP 3: PatientID 매칭 → 신장/체중/BMI/SMI 기입 ─────────────────────────
# 각 환자의 코호트 기록을 최신순으로 순회하며
# 신장·체중이 모두 유효한(non-null, 범위 내) 첫 번째 기록을 사용한다.
# BMI와 SMI도 같은 기록에서 계산해 한 번에 기입한다.

print("\n3. VB 코호트 매칭 및 지표 기입 중...")

for r_idx in tqdm(range(header_row_idx + 1, ws.max_row + 1), desc="Mapping & Calculating"):
    pid_val = ws.cell(row=r_idx, column=pid_col).value
    if pid_val is None or str(pid_val).strip() == "None":
        continue

    # PatientID에 ".0" 소수점이 붙어 있을 경우 정수 부분만 추출
    pid_key = str(pid_val).strip().split('.')[0]

    # TAMA·IMATA 값 읽기 (SMI 계산에 필요)
    raw_tama  = ws.cell(row=r_idx, column=tama_col).value
    raw_imata = ws.cell(row=r_idx, column=imata_col).value
    tama  = float(raw_tama)  if raw_tama  is not None else 0.0
    imata = float(raw_imata) if raw_imata is not None else 0.0

    # 코호트 딕셔너리에 없는 환자는 건너뜀
    if pid_key not in cohort_dict:
        continue

    # 최신순 기록을 순회하며 신장·체중이 모두 유효한 기록 탐색
    height, weight = None, None
    for record in cohort_dict[pid_key]:
        # '신장(cm)' 또는 '신장' 컬럼 중 존재하는 것 사용
        h = record.get('신장(cm)') or record.get('신장')
        w = record.get('체중(kg)') or record.get('체중')
        # 결측값 건너뜀
        if h is None or pd.isna(h) or w is None or pd.isna(w):
            continue
        # 이상치 범위 건너뜀
        if not in_range(h, VALID_HEIGHT) or not in_range(w, VALID_WEIGHT):
            continue
        # 신장 단위 자동 변환: 3.0 이하면 미터(m) 단위로 간주
        h_m = float(h) / 100.0 if float(h) > 3.0 else float(h)
        # BMI가 유효 범위인지 확인
        if h_m <= 0 or not in_range(float(w) / (h_m ** 2), VALID_BMI):
            continue
        height, weight = h, w
        break  # 유효한 기록 발견 → 최신 기록 사용

    # 유효한 신장·체중을 셀에 기입
    if height is not None and height_col is not None:
        ws.cell(row=r_idx, column=height_col).value = float(height)
    if weight is not None and weight_col is not None:
        ws.cell(row=r_idx, column=weight_col).value = float(weight)

    # BMI 계산: weight / height_m²
    if height is not None and weight is not None and bmi_col is not None:
        height_m = float(height) / 100.0 if float(height) > 3.0 else float(height)
        if height_m > 0:
            ws.cell(row=r_idx, column=bmi_col).value = round(
                float(weight) / (height_m ** 2), 2
            )

    # SMI 계산: (TAMA - IMATA) / height_m²
    # TAMA에서 지방(IMATA)을 제거한 순수 근육 단면적을 키로 보정
    if height is not None and smi_col is not None:
        try:
            height_m = float(height) / 100.0 if float(height) > 3.0 else float(height)
            if height_m > 0:
                ws.cell(row=r_idx, column=smi_col).value = round(
                    (tama - imata) / (height_m ** 2), 2
                )
        except ZeroDivisionError:
            pass


# ── STEP 4: BMI 결측 보완 ─────────────────────────────────────────────────────
# STEP 3에서 코호트 매칭에 실패했지만 Height·Weight가 이미 채워진 행에 대해
# BMI를 재계산한다 (기존 엑셀에 Height·Weight가 미리 입력된 경우 처리).

print("\n4. BMI 결측치 보완 중...")
bmi_filled = 0
for r_idx in range(header_row_idx + 1, ws.max_row + 1):
    if bmi_col is None or height_col is None or weight_col is None:
        break
    if ws.cell(row=r_idx, column=bmi_col).value is not None:
        continue   # 이미 값이 있으면 건너뜀
    height_val = ws.cell(row=r_idx, column=height_col).value
    weight_val = ws.cell(row=r_idx, column=weight_col).value
    if height_val is None or weight_val is None:
        continue
    try:
        h = float(height_val)
        w = float(weight_val)
        height_m = h / 100.0 if h > 3.0 else h
        if height_m > 0:
            ws.cell(row=r_idx, column=bmi_col).value = round(w / (height_m ** 2), 2)
            bmi_filled += 1
    except (ValueError, ZeroDivisionError):
        pass
print(f"-> BMI 보완 완료: {bmi_filled}명")


# ── STEP 5: IMATA 결측 행 제거 ────────────────────────────────────────────────
# IMATA가 없으면 SMI를 계산할 수 없으므로 해당 행을 제거한다.
# 역순 순회 이유: delete_rows()가 이후 행의 인덱스를 당기기 때문.

print("\n5. IMATA 결측 행 제거 중...")
removed = 0
for r_idx in range(ws.max_row, header_row_idx, -1):
    val = ws.cell(row=r_idx, column=imata_col).value
    if val is None or str(val).strip() in ("", "None"):
        ws.delete_rows(r_idx)
        removed += 1
print(f"-> {removed}행 제거 완료")


# ── STEP 6: Height 또는 Weight 결측 행 제거 ──────────────────────────────────
# 신장·체중이 없으면 BMI·SMI 산출이 불가능하므로 해당 행을 제거한다.

print("\n6. Height·Weight 결측 행 제거 중...")
removed_hw = 0
for r_idx in range(ws.max_row, header_row_idx, -1):
    h_val = ws.cell(row=r_idx, column=height_col).value if height_col else None
    w_val = ws.cell(row=r_idx, column=weight_col).value if weight_col else None
    if h_val is None or w_val is None:
        ws.delete_rows(r_idx)
        removed_hw += 1
print(f"-> {removed_hw}행 제거 완료")


# ── STEP 7: SMI 결측 보완 ─────────────────────────────────────────────────────
# STEP 3에서 SMI가 채워지지 않은 행에 대해 재계산을 시도한다.
# Height·TAMA·IMATA 가 모두 있어야 계산 가능하다.

print("\n7. SMI 결측치 보완 중...")
smi_filled = 0
for r_idx in range(header_row_idx + 1, ws.max_row + 1):
    if smi_col is None or height_col is None:
        break
    if ws.cell(row=r_idx, column=smi_col).value is not None:
        continue   # 이미 값이 있으면 건너뜀
    height_val = ws.cell(row=r_idx, column=height_col).value
    tama_val   = ws.cell(row=r_idx, column=tama_col).value
    imata_val  = ws.cell(row=r_idx, column=imata_col).value
    if height_val is None or tama_val is None or imata_val is None:
        continue
    try:
        h = float(height_val)
        height_m = h / 100.0 if h > 3.0 else h
        if height_m > 0:
            ws.cell(row=r_idx, column=smi_col).value = round(
                (float(tama_val) - float(imata_val)) / (height_m ** 2), 2
            )
            smi_filled += 1
    except (ValueError, ZeroDivisionError):
        pass
print(f"-> SMI 보완 완료: {smi_filled}명")


# ── STEP 8: 이상치 행 삭제 ────────────────────────────────────────────────────
# 4개 지표(신장·체중·BMI·SMI) 중 하나라도 유효 범위를 벗어난 행을 제거한다.
# 하나의 값이라도 이상치이면 해당 행 전체를 삭제해 데이터 신뢰성을 보장한다.

print("\n8. 이상치 행 삭제 중...")
removed_outlier = 0
for r_idx in range(ws.max_row, header_row_idx, -1):
    h_val   = ws.cell(row=r_idx, column=height_col).value if height_col else None
    w_val   = ws.cell(row=r_idx, column=weight_col).value if weight_col else None
    bmi_val = ws.cell(row=r_idx, column=bmi_col).value    if bmi_col    else None
    smi_val = ws.cell(row=r_idx, column=smi_col).value    if smi_col    else None

    # 각 값이 존재하면서 범위를 벗어나면 이상치로 판정
    is_outlier = (
        (h_val   is not None and not in_range(h_val,   VALID_HEIGHT)) or
        (w_val   is not None and not in_range(w_val,   VALID_WEIGHT)) or
        (bmi_val is not None and not in_range(bmi_val, VALID_BMI))    or
        (smi_val is not None and not in_range(smi_val, VALID_SMI))
    )
    if is_outlier:
        ws.delete_rows(r_idx)
        removed_outlier += 1
print(f"-> 이상치 {removed_outlier}행 삭제 완료")


# ── STEP 9: 최종 저장 ─────────────────────────────────────────────────────────

print(f"\n9. 최종 파일 저장 중...")
wb.save(final_output_path)
wb.close()

print(f"\n[완료] SMI 산출 결과 저장: '{final_output_path}'")
