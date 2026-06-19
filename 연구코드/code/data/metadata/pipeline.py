"""
[전체 목적]
PHASE 0: DLO 결과 엑셀 중복 제거 + OCR 메타데이터 보완
  → {SITE}_DLO_Results.xlsx → {SITE}_DLO_Results_Unique.xlsx

PHASE 1: VB 코호트 신장·체중 매핑 → BMI·SMI 계산·정제
  → {SITE}_DLO_Results_Unique.xlsx → {SITE}_DLO_Results_SMI.xlsx

[핵심 수식]
  BMI = 체중(kg) / 신장(m)²
  SMI = (TAMA - IMATA) / 신장(m)²   [단위: cm²/m²]
"""

import os
import re
import cv2
import numpy as np
import pandas as pd
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from tqdm import tqdm

# ── 공통 환경 설정 ────────────────────────────────────────────────────────────
# 여기만 바꾸면 전체 경로·코드가 자동으로 연동된다.

SITE = "신촌"   # "강남" 또는 "신촌"

SITE_CODE_MAP = {"강남": "20", "신촌": "10"}
SITE_CODE = SITE_CODE_MAP[SITE]

BASE_DIR          = rf"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\{SITE}\metadata"
DLO_BASE          = rf"D:\영상제공\{SITE}\{SITE}_결과"
file_path         = os.path.join(BASE_DIR, f"{SITE}_DLO_Results.xlsx")
output_path       = os.path.join(BASE_DIR, f"{SITE}_DLO_Results_Unique.xlsx")
vb_cohort_path    = r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\radiation_data_260421_VBcohort.xlsx"
final_output_path = os.path.join(BASE_DIR, f"{SITE}_DLO_Results_SMI.xlsx")

SAVE_INTERVAL = 50

VALID_HEIGHT = (130.0, 210.0)
VALID_WEIGHT = (25.0,  200.0)
VALID_BMI    = (10.0,  55.0)
VALID_SMI    = (10.0,  80.0)

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
_SERIES_PAT = re.compile(r"Series\s*Desc[^\n]*?:\s*([^:\n]+)", re.IGNORECASE)


# ── 공통 함수 ─────────────────────────────────────────────────────────────────

def parse_dlo_img_path(formula_str, dlo_base):
    if not formula_str:
        return None
    m = re.search(r'HYPERLINK\("([^"]+)"', str(formula_str))
    if not m:
        return None
    rel  = m.group(1).replace("\\", "/").lstrip("./")
    full = os.path.join(dlo_base, rel.replace("/", os.sep))
    return full


def extract_metadata_from_sarco(img_path):
    if not img_path or not os.path.exists(img_path):
        return {}
    results = {}
    try:
        img = Image.open(img_path)
        w, h = img.size
        arr  = np.array(img)
        gray = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
        _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)

        full_text   = pytesseract.image_to_string(thresh, config="--psm 3")
        model_match = re.search(
            r"(?:Manufacturer\s*Model\s*Name|Model\s*Name)\s*[:|-]?\s*([^\n]+)",
            full_text, re.IGNORECASE
        )
        if model_match:
            results["ManufacturerModelName"] = model_match.group(1).strip()

        series_match = _SERIES_PAT.search(full_text)
        if series_match:
            val = series_match.group(1).strip()
            val = re.sub(r'\s+[A-Z][a-zA-Z\s/]*$', '', val).strip()
            results["SeriesDescription"] = val

        crop      = img.crop((0, int(h * 0.72), w, int(h * 0.93)))
        crop_arr  = np.array(crop)
        crop_gray = cv2.cvtColor(crop_arr, cv2.COLOR_RGB2GRAY)
        _, crop_thresh = cv2.threshold(crop_gray, 180, 255, cv2.THRESH_BINARY)
        table_text = pytesseract.image_to_string(crop_thresh, config="--psm 6")

        for line in table_text.split("\n"):
            line = line.strip()
            m = re.match(r"(TAMA|IMATA)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)", line)
            if m:
                vals = [int(m.group(i)) for i in range(2, 6)]
                results[m.group(1)] = vals[3]

        return results
    except Exception:
        return results


def in_range(value, bounds: tuple) -> bool:
    try:
        return bounds[0] <= float(value) <= bounds[1]
    except (TypeError, ValueError):
        return False


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 0: 중복 제거 + OCR 메타데이터 보완
# ══════════════════════════════════════════════════════════════════════════════

print("=" * 78)
print(f"PHASE 0 [{SITE}]: 중복 제거 및 OCR 메타데이터 보완")
print("=" * 78)

# STEP 0-1: 엑셀 로드 및 컬럼 위치 탐색
print("\n[0-1] 엑셀 파일 로드 및 구조 분석 중...")
wb = load_workbook(file_path)
ws = wb.active

col_indices = {
    "PatientID": None, "SRC_Report": None,
    "TAMA": None, "IMATA": None,
    "ManufacturerModelName": None, "SeriesDescription": None,
}
header_row_idx = None

for r_idx in range(1, 11):
    for c_idx in range(1, ws.max_column + 1):
        val = str(ws.cell(row=r_idx, column=c_idx).value).strip()
        if val in col_indices:
            col_indices[val] = c_idx
            header_row_idx   = r_idx
    if col_indices["PatientID"] and col_indices["SRC_Report"]:
        break

if not col_indices["PatientID"] or not col_indices["SRC_Report"]:
    raise ValueError("필수 컬럼('PatientID' 또는 'SRC_Report')을 찾을 수 없습니다.")

for col_name in ("ManufacturerModelName", "SeriesDescription"):
    if col_indices[col_name] is None:
        new_col = ws.max_column + 1
        ws.cell(row=header_row_idx, column=new_col).value = col_name
        col_indices[col_name] = new_col
        print(f"-> '{col_name}' 컬럼 없음 → {new_col}번째 열에 추가")

pid_col    = col_indices["PatientID"]
src_col    = col_indices["SRC_Report"]
tama_col   = col_indices["TAMA"]
imata_col  = col_indices["IMATA"]
model_col  = col_indices["ManufacturerModelName"]
series_col = col_indices["SeriesDescription"]
start_row  = header_row_idx + 1
max_row    = ws.max_row

# STEP 0-2: PatientID 정제 및 중복·비숫자 행 역순 제거
print("\n[0-2] PatientID 정제 및 중복 행 제거 중...")
row_info_list = []

for r_idx in range(start_row, max_row + 1):
    pid_cell = ws.cell(row=r_idx, column=pid_col)
    raw_pid  = str(pid_cell.value).strip() if pid_cell.value is not None else ""
    if not raw_pid or raw_pid == "None":
        row_info_list.append((r_idx, "None", False))
        continue
    try:
        pid_int        = int(float(raw_pid))
        pid_cell.value = pid_int
        pid_str_key    = str(pid_int)
    except (ValueError, TypeError):
        pid_str_key = raw_pid

    is_tama_numeric = False
    if tama_col:
        tama_val = ws.cell(row=r_idx, column=tama_col).value
        if tama_val is not None:
            try:
                float(str(tama_val).strip())
                is_tama_numeric = True
            except ValueError:
                pass
    row_info_list.append((r_idx, pid_str_key, is_tama_numeric))

seen_pids = set()
removed_dup = removed_tama = removed_none = 0
for r_idx, pid_key, is_tama_numeric in tqdm(
    reversed(row_info_list), total=len(row_info_list), desc="Cleaning Rows"
):
    if pid_key == "None":
        ws.delete_rows(r_idx); removed_none += 1
    elif tama_col and not is_tama_numeric:
        ws.delete_rows(r_idx); removed_tama += 1
    elif pid_key in seen_pids:
        ws.delete_rows(r_idx); removed_dup += 1
    else:
        seen_pids.add(pid_key)

updated_max_row = ws.max_row
print(f"\n[0-2] 완료 — 유니크 {updated_max_row - start_row + 1}명 "
      f"(중복 {removed_dup}, TAMA오류 {removed_tama}, 빈행 {removed_none} 제거)")

# STEP 0-3: 유니크 환자별 OCR 매핑
print(f"\n[0-3] OCR 분석 중... (매 {SAVE_INTERVAL}건마다 자동 저장)")
processed = ocr_log = 0
for r_idx in tqdm(range(start_row, updated_max_row + 1), desc="OCR & Data Mapping"):
    pid_cell = ws.cell(row=r_idx, column=pid_col)
    if pid_cell.value is None or str(pid_cell.value).strip() == "None":
        continue
    img_path = parse_dlo_img_path(ws.cell(row=r_idx, column=src_col).value, DLO_BASE)
    if img_path and os.path.exists(img_path):
        ocr = extract_metadata_from_sarco(img_path)
        if tama_col   and "TAMA"  in ocr and ocr["TAMA"]  != 0:
            ws.cell(row=r_idx, column=tama_col).value   = int(ocr["TAMA"])
        if imata_col  and "IMATA" in ocr and ocr["IMATA"] != 0:
            ws.cell(row=r_idx, column=imata_col).value  = int(ocr["IMATA"])
        if model_col  and "ManufacturerModelName" in ocr:
            ws.cell(row=r_idx, column=model_col).value  = str(ocr["ManufacturerModelName"])
        if series_col and "SeriesDescription" in ocr:
            ws.cell(row=r_idx, column=series_col).value = str(ocr["SeriesDescription"])
        processed += 1
        if ocr_log < 5:
            tqdm.write(f"   [OCR] PatientID={pid_cell.value} "
                       f"TAMA={ocr.get('TAMA', 0)} IMATA={ocr.get('IMATA', 0)}")
            ocr_log += 1
        if processed % SAVE_INTERVAL == 0:
            wb.save(output_path)
            tqdm.write(f"   [자동 저장] {processed}건 완료")
    elif img_path and ocr_log < 5:
        tqdm.write(f"   [경고] 이미지 없음: {os.path.basename(img_path)} "
                   f"(PatientID={pid_cell.value})")

# STEP 0-4: 최종 저장
print(f"\n[0-4] 최종 저장 중...")
wb.save(output_path)
wb.close()
print(f"[PHASE 0 완료] '{output_path}'")


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1: VB 코호트 매핑 → BMI·SMI 계산·정제
# ══════════════════════════════════════════════════════════════════════════════

print("\n" + "=" * 78)
print(f"PHASE 1 [{SITE}]: VB 코호트 매핑 및 BMI·SMI 계산 (병원코드: {SITE_CODE})")
print("=" * 78)

# STEP 1-1: VB 코호트 로드
print("\n[1-1] VB 코호트 로드 중...")
try:
    df_cohort = pd.read_excel(vb_cohort_path, sheet_name='신장체중(예진)_전체', dtype=str)
    df_cohort = df_cohort[df_cohort['지역병원코드'] == SITE_CODE].copy()
    df_cohort['신장'] = pd.to_numeric(df_cohort['신장'], errors='coerce')
    df_cohort['체중'] = pd.to_numeric(df_cohort['체중'], errors='coerce')
    df_cohort['PatientID_key'] = df_cohort['연구등록번호'].str.strip().str.split('.').str[0]
    df_cohort = df_cohort.sort_values('내원연월', ascending=False)
    cohort_dict = {
        pid: records.to_dict('records')
        for pid, records in df_cohort.groupby('PatientID_key', sort=False)
    }
    print(f"-> {SITE} 예진 데이터 {len(cohort_dict)}명 확보.")
except Exception as e:
    raise IOError(f"VBcohort.xlsx 파일을 읽는 중 오류가 발생했습니다: {e}")

# STEP 1-2: Unique 엑셀 로드 및 컬럼 구조 파악
# PHASE 0이 저장한 output_path를 그대로 입력으로 사용한다.
print("\n[1-2] Unique 엑셀 로드 및 구조 분석 중...")
wb = load_workbook(output_path)
ws = wb.active
assert ws is not None, "활성 시트를 찾을 수 없습니다."

col_indices = {
    "PatientID": None, "TAMA": None, "IMATA": None,
    "BMI": None, "Height": None, "Weight": None, "SMI": None,
}
header_row_idx = None
max_col: int = ws.max_column or 1

for r_idx in range(1, 11):
    for c_idx in range(1, max_col + 1):
        val = str(ws.cell(row=r_idx, column=c_idx).value).strip()
        if val in col_indices:
            col_indices[val] = c_idx
            header_row_idx   = r_idx
    if col_indices["PatientID"] and col_indices["TAMA"]:
        break

if not col_indices["PatientID"] or header_row_idx is None:
    raise ValueError("시트 안에서 'PatientID' 컬럼을 찾을 수 없습니다.")

new_columns = ["Height", "Weight", "SMI"]
if col_indices["BMI"] is None:
    new_columns.insert(0, "BMI")

current_max_col: int = ws.max_column or max_col
for col_name in new_columns:
    if col_indices[col_name] is None:
        current_max_col += 1
        ws.cell(row=header_row_idx, column=current_max_col).value = col_name
        col_indices[col_name] = current_max_col
        print(f"-> '{col_name}' 컬럼 추가 ({current_max_col}번째 열)")

pid_col    = col_indices["PatientID"]
tama_col   = col_indices["TAMA"]
imata_col  = col_indices["IMATA"]
bmi_col    = col_indices["BMI"]
height_col = col_indices["Height"]
weight_col = col_indices["Weight"]
smi_col    = col_indices["SMI"]
assert pid_col is not None and tama_col is not None and imata_col is not None

# STEP 1-3: PatientID 매칭 → 신장/체중/BMI/SMI 기입
print("\n[1-3] VB 코호트 매칭 및 지표 기입 중...")
for r_idx in tqdm(range(header_row_idx + 1, ws.max_row + 1), desc="Mapping & Calculating"):
    pid_val = ws.cell(row=r_idx, column=pid_col).value
    if pid_val is None or str(pid_val).strip() == "None":
        continue
    pid_key   = str(pid_val).strip().split('.')[0]
    raw_tama  = ws.cell(row=r_idx, column=tama_col).value
    raw_imata = ws.cell(row=r_idx, column=imata_col).value
    tama  = float(raw_tama)  if raw_tama  is not None else 0.0
    imata = float(raw_imata) if raw_imata is not None else 0.0

    if pid_key not in cohort_dict:
        continue

    height, weight = None, None
    for record in cohort_dict[pid_key]:
        h = record.get('신장(cm)') or record.get('신장')
        w = record.get('체중(kg)') or record.get('체중')
        if h is None or pd.isna(h) or w is None or pd.isna(w):
            continue
        if not in_range(h, VALID_HEIGHT) or not in_range(w, VALID_WEIGHT):
            continue
        h_m = float(h) / 100.0 if float(h) > 3.0 else float(h)
        if h_m <= 0 or not in_range(float(w) / (h_m ** 2), VALID_BMI):
            continue
        height, weight = h, w
        break

    if height is not None and height_col:
        ws.cell(row=r_idx, column=height_col).value = float(height)
    if weight is not None and weight_col:
        ws.cell(row=r_idx, column=weight_col).value = float(weight)
    if height is not None and weight is not None and bmi_col:
        h_m = float(height) / 100.0 if float(height) > 3.0 else float(height)
        if h_m > 0:
            ws.cell(row=r_idx, column=bmi_col).value = round(float(weight) / (h_m ** 2), 2)
    if height is not None and smi_col:
        try:
            h_m = float(height) / 100.0 if float(height) > 3.0 else float(height)
            if h_m > 0:
                ws.cell(row=r_idx, column=smi_col).value = round((tama - imata) / (h_m ** 2), 2)
        except ZeroDivisionError:
            pass

# STEP 1-4: BMI 결측 보완
print("\n[1-4] BMI 결측치 보완 중...")
bmi_filled = 0
for r_idx in range(header_row_idx + 1, ws.max_row + 1):
    if not (bmi_col and height_col and weight_col):
        break
    if ws.cell(row=r_idx, column=bmi_col).value is not None:
        continue
    h_val = ws.cell(row=r_idx, column=height_col).value
    w_val = ws.cell(row=r_idx, column=weight_col).value
    if h_val is None or w_val is None:
        continue
    try:
        h = float(h_val)
        h_m = h / 100.0 if h > 3.0 else h
        if h_m > 0:
            ws.cell(row=r_idx, column=bmi_col).value = round(float(w_val) / (h_m ** 2), 2)
            bmi_filled += 1
    except (ValueError, ZeroDivisionError):
        pass
print(f"-> BMI 보완 완료: {bmi_filled}명")

# STEP 1-5: IMATA 결측 행 제거
print("\n[1-5] IMATA 결측 행 제거 중...")
removed = 0
for r_idx in range(ws.max_row, header_row_idx, -1):
    val = ws.cell(row=r_idx, column=imata_col).value
    if val is None or str(val).strip() in ("", "None"):
        ws.delete_rows(r_idx); removed += 1
print(f"-> {removed}행 제거 완료")

# STEP 1-6: Height·Weight 결측 행 제거
print("\n[1-6] Height·Weight 결측 행 제거 중...")
removed_hw = 0
for r_idx in range(ws.max_row, header_row_idx, -1):
    h_val = ws.cell(row=r_idx, column=height_col).value if height_col else None
    w_val = ws.cell(row=r_idx, column=weight_col).value if weight_col else None
    if h_val is None or w_val is None:
        ws.delete_rows(r_idx); removed_hw += 1
print(f"-> {removed_hw}행 제거 완료")

# STEP 1-7: SMI 결측 보완
print("\n[1-7] SMI 결측치 보완 중...")
smi_filled = 0
for r_idx in range(header_row_idx + 1, ws.max_row + 1):
    if not (smi_col and height_col):
        break
    if ws.cell(row=r_idx, column=smi_col).value is not None:
        continue
    h_val  = ws.cell(row=r_idx, column=height_col).value
    t_val  = ws.cell(row=r_idx, column=tama_col).value
    im_val = ws.cell(row=r_idx, column=imata_col).value
    if h_val is None or t_val is None or im_val is None:
        continue
    try:
        h = float(h_val)
        h_m = h / 100.0 if h > 3.0 else h
        if h_m > 0:
            ws.cell(row=r_idx, column=smi_col).value = round(
                (float(t_val) - float(im_val)) / (h_m ** 2), 2
            )
            smi_filled += 1
    except (ValueError, ZeroDivisionError):
        pass
print(f"-> SMI 보완 완료: {smi_filled}명")

# STEP 1-8: 이상치 행 삭제
print("\n[1-8] 이상치 행 삭제 중...")
removed_outlier = 0
for r_idx in range(ws.max_row, header_row_idx, -1):
    h_val   = ws.cell(row=r_idx, column=height_col).value if height_col else None
    w_val   = ws.cell(row=r_idx, column=weight_col).value if weight_col else None
    bmi_val = ws.cell(row=r_idx, column=bmi_col).value    if bmi_col    else None
    smi_val = ws.cell(row=r_idx, column=smi_col).value    if smi_col    else None
    is_outlier = (
        (h_val   is not None and not in_range(h_val,   VALID_HEIGHT)) or
        (w_val   is not None and not in_range(w_val,   VALID_WEIGHT)) or
        (bmi_val is not None and not in_range(bmi_val, VALID_BMI))    or
        (smi_val is not None and not in_range(smi_val, VALID_SMI))
    )
    if is_outlier:
        ws.delete_rows(r_idx); removed_outlier += 1
print(f"-> 이상치 {removed_outlier}행 삭제 완료")

# STEP 1-9: 최종 저장
print(f"\n[1-9] 최종 파일 저장 중...")
wb.save(final_output_path)
wb.close()
print(f"[PHASE 1 완료] '{final_output_path}'")

print("\n" + "=" * 78)
print(f"[전체 파이프라인 완료] {SITE} 데이터 처리 종료")
print("=" * 78)
