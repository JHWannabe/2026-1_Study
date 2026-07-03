"""
강남_DLO_Results_Unique.xlsx - 'kVp_100_조영제X_제거' 시트
중복 PatientID 행 중 남길 행을 GUI로 선택 후 저장
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font
import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path

EXCEL_PATH = Path(r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\강남\metadata\강남_DLO_Results_Unique.xlsx")
SHEET_NAME = "kVp_100"

# 숫자형으로 보여줄 컬럼 소수점 처리
FLOAT_COLS = {"TAMA", "IMATA", "PatientAge"}


def load_data():
    df = pd.read_excel(EXCEL_PATH, sheet_name=SHEET_NAME)
    df = df.reset_index(drop=True)
    return df


def load_hyperlinks(sheet_name: str) -> dict[int, dict]:
    """원본 시트의 SRC_Report 하이퍼링크/수식을 0-based 행 인덱스로 반환.

    네이티브 hyperlink 객체와 =HYPERLINK() 수식 문자열 모두 처리.
    """
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    src_col = headers.index("SRC_Report")  # 0-based
    links = {}
    for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
        cell = row[src_col]
        entry: dict = {}
        if cell.hyperlink:
            entry["hyperlink"] = cell.hyperlink
        if cell.value is not None and "HYPERLINK" in str(cell.value).upper():
            entry["formula"] = str(cell.value)
        if entry:
            links[row_idx] = entry
    wb.close()
    return links


def get_dup_groups(df):
    mask = df.duplicated("PatientID", keep=False)
    dup_df = df[mask].copy()
    groups = {pid: grp for pid, grp in dup_df.groupby("PatientID")}
    return groups  # dict[PatientID -> DataFrame]


class DupResolverApp(tk.Tk):
    def __init__(self, df, groups):
        super().__init__()
        self.df = df
        self.groups = groups
        self.pids = sorted(groups.keys())
        self.selections = {}   # PatientID -> 선택된 원본 df index
        self.cur = 0           # 현재 그룹 포인터

        self.hyperlinks = load_hyperlinks(SHEET_NAME)  # {orig_idx: hyperlink}

        self.title("중복 PatientID 선택")
        self.geometry("1100x600")
        self.resizable(True, True)

        self._build_ui()
        self._load_group(0)

    # ── UI 빌드 ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        top = tk.Frame(self)
        top.pack(fill="x", padx=10, pady=(8, 0))

        self.lbl_progress = tk.Label(top, text="", font=("Arial", 11, "bold"))
        self.lbl_progress.pack(side="left")

        self.lbl_pid = tk.Label(top, text="", font=("Arial", 11))
        self.lbl_pid.pack(side="left", padx=20)

        # ── 테이블 영역 ──
        mid = tk.Frame(self, bd=1, relief="sunken")
        mid.pack(fill="both", expand=True, padx=10, pady=8)

        # 라디오 선택 열 + 데이터 열
        self.cols = ["선택", "원본행"] + list(self.df.columns)
        self.tree = ttk.Treeview(mid, columns=self.cols, show="headings", height=6)

        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        mid.rowconfigure(0, weight=1)
        mid.columnconfigure(0, weight=1)

        col_widths = {"선택": 50, "원본행": 60, "PatientID": 90, "PatientAge": 80,
                      "PatientSex": 70, "kVp": 50, "SRC_Report": 80,
                      "TAMA": 60, "IMATA": 60, "Manufacturer": 200, "Series_Desc": 250}
        for c in self.cols:
            w = col_widths.get(c, 100)
            self.tree.heading(c, text=c)
            self.tree.column(c, width=w, minwidth=40, anchor="center")

        self.tree.bind("<ButtonRelease-1>", self._on_click)

        # ── 하단 버튼 ──
        bot = tk.Frame(self)
        bot.pack(fill="x", padx=10, pady=(0, 10))

        self.btn_prev = tk.Button(bot, text="◀ 이전", width=10, command=self._prev)
        self.btn_prev.pack(side="left", padx=4)

        self.btn_next = tk.Button(bot, text="다음 ▶", width=10, command=self._next)
        self.btn_next.pack(side="left", padx=4)

        tk.Label(bot, text="  ").pack(side="left")

        self.btn_auto = tk.Button(bot, text="Portal 우선 자동선택", width=20,
                                  command=self._auto_select_all)
        self.btn_auto.pack(side="left", padx=4)

        self.btn_save = tk.Button(bot, text="저장 (새 시트)", width=16,
                                  bg="#2196F3", fg="white", font=("Arial", 10, "bold"),
                                  command=self._save)
        self.btn_save.pack(side="right", padx=4)

        self.lbl_status = tk.Label(bot, text="", fg="gray")
        self.lbl_status.pack(side="right", padx=10)

    # ── 그룹 로드 ────────────────────────────────────────────────────────────

    def _load_group(self, idx):
        self.cur = idx
        pid = self.pids[idx]
        grp = self.groups[pid]

        self.lbl_progress.config(
            text=f"[{idx + 1} / {len(self.pids)}]"
        )
        self.lbl_pid.config(text=f"PatientID: {pid}  ({len(grp)}행)")

        self.tree.delete(*self.tree.get_children())
        selected_orig = self.selections.get(pid)

        for orig_idx, row in grp.iterrows():
            mark = "★" if orig_idx == selected_orig else ""
            vals = [mark, orig_idx] + [
                (int(v) if col in FLOAT_COLS and pd.notna(v) else v)
                for col, v in zip(self.df.columns, row)
            ]
            iid = self.tree.insert("", "end", values=vals, tags=(str(orig_idx),))
            if orig_idx == selected_orig:
                self.tree.item(iid, tags=("selected",))
                self.tree.tag_configure("selected", background="#BBDEFB")

        self.btn_prev.config(state="normal" if idx > 0 else "disabled")
        self.btn_next.config(state="normal" if idx < len(self.pids) - 1 else "disabled")

        unresolved = sum(1 for p in self.pids if p not in self.selections)
        self.lbl_status.config(text=f"미결정: {unresolved}건")

    # ── 클릭으로 행 선택 ─────────────────────────────────────────────────────

    def _on_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return
        vals = self.tree.item(item, "values")
        orig_idx = int(vals[1])
        pid = self.pids[self.cur]
        self.selections[pid] = orig_idx
        self._load_group(self.cur)

    # ── 이전/다음 ────────────────────────────────────────────────────────────

    def _prev(self):
        if self.cur > 0:
            self._load_group(self.cur - 1)

    def _next(self):
        if self.cur < len(self.pids) - 1:
            self._load_group(self.cur + 1)

    # ── 자동 선택 (Portal > Delay > Arterial 우선) ───────────────────────────

    PRIORITY_KEYWORDS = ["portal", "por", "with", "post", "delay", "contrast", "arterial", "artery"]

    def _auto_select_group(self, pid):
        grp = self.groups[pid]
        for kw in self.PRIORITY_KEYWORDS:
            for orig_idx, row in grp.iterrows():
                if kw in str(row.get("Series_Desc", "")).lower():
                    return orig_idx
        return grp.index[0]  # fallback: 첫 번째 행

    def _auto_select_all(self):
        count = 0
        for pid in self.pids:
            if pid not in self.selections:
                self.selections[pid] = self._auto_select_group(pid)
                count += 1
        self._load_group(self.cur)
        messagebox.showinfo("자동선택 완료", f"{count}건 자동 선택 완료.\n결과를 검토한 후 저장하세요.")

    # ── 저장 ─────────────────────────────────────────────────────────────────

    def _save(self):
        unresolved = [p for p in self.pids if p not in self.selections]
        if unresolved:
            ans = messagebox.askyesno(
                "미결정 항목 있음",
                f"아직 {len(unresolved)}개 PatientID가 선택되지 않았습니다.\n"
                "미결정 항목은 자동선택(Portal 우선)으로 처리할까요?"
            )
            if ans:
                for pid in unresolved:
                    self.selections[pid] = self._auto_select_group(pid)
            else:
                return

        # 남길 행 인덱스: 비중복 행 + 선택된 중복 행
        dup_mask = self.df.duplicated("PatientID", keep=False)
        keep_from_dup = set(self.selections.values())
        keep_mask = (~dup_mask) | (self.df.index.isin(keep_from_dup))
        filtered = self.df[keep_mask]           # 원본 index 보존
        orig_indices = filtered.index.tolist()  # 하이퍼링크 복원에 사용
        result_df = filtered.reset_index(drop=True)

        removed = len(self.df) - len(result_df)
        out_sheet = "kVp_100_조영제X_중복제거"

        # pandas로 데이터 저장
        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a",
                            if_sheet_exists="replace") as writer:
            result_df.to_excel(writer, sheet_name=out_sheet, index=False)

        # SRC_Report 하이퍼링크 복원
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb[out_sheet]
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        src_col = headers.index("SRC_Report") + 1  # 1-based
        for new_row, orig_idx in enumerate(orig_indices, start=2):  # 헤더=1행, 데이터=2행~
            entry = self.hyperlinks.get(orig_idx)
            if entry:
                cell = ws.cell(row=new_row, column=src_col)
                if "hyperlink" in entry:
                    cell.hyperlink = entry["hyperlink"]
                if "formula" in entry:
                    cell.value = entry["formula"]
                cell.font = Font(color="0563C1", underline="single")
        wb.save(EXCEL_PATH)

        messagebox.showinfo(
            "저장 완료",
            f"시트 '{out_sheet}'에 저장되었습니다.\n"
            f"원본: {len(self.df)}행  →  저장: {len(result_df)}행  (제거: {removed}행)"
        )
        self.lbl_status.config(text=f"저장 완료 — {len(result_df)}행")


def main():
    df = load_data()
    groups = get_dup_groups(df)
    print(f"전체 행: {len(df)},  중복 PatientID: {len(groups)}개")
    app = DupResolverApp(df, groups)
    app.mainloop()


if __name__ == "__main__":
    main()
