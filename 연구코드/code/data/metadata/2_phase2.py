"""
[목적]
PHASE 1: 강남 VB 코호트 신장·체중 매핑 → BMI·SMI 계산·정제
  입력 시트: 강남_DLO_Results_Unique.xlsx / "kVp_100_조영제X_중복제거"
  출력: 강남_DLO_Results_SMI_kVp100.xlsx

[핵심 수식]
  BMI = 체중(kg) / 신장(m)²
  SMI = (TAMA - IMATA) / 신장(m)²   [단위: cm²/m²]
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font
from tqdm import tqdm

# ── 환경 설정 ─────────────────────────────────────────────────────────────────

SITE      = "강남"
SITE_CODE = "20"

BASE_DIR          = rf"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\{SITE}\metadata"
INPUT_PATH        = os.path.join(BASE_DIR, f"{SITE}_DLO_Results_Unique.xlsx")
INPUT_SHEET       = "kVp_100"
VB_COHORT_PATH    = r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\radiation_data_260421_VBcohort.xlsx"
FINAL_OUTPUT_PATH = os.path.join(BASE_DIR, f"{SITE}_DLO_Results_SMI_kVp100.xlsx")

VALID_HEIGHT = (130.0, 210.0)
VALID_WEIGHT = (25.0,  200.0)
VALID_BMI    = (10.0,  55.0)

_BOLD_CENTER = Font(bold=True), Alignment(horizontal="center")


# ── 공통 함수 ─────────────────────────────────────────────────────────────────

def wcell(ws, row: int, col: int) -> Cell:
    return ws.cell(row=row, column=col)  # type: ignore[return-value]


def in_range(value, bounds: tuple) -> bool:
    try:
        return bounds[0] <= float(value) <= bounds[1]
    except (TypeError, ValueError):
        return False


def drop_blank_columns(ws, header_row: int | None):
    if header_row is None:
        return
    for c_idx in range(ws.max_column, 0, -1):
        if ws.cell(row=header_row, column=c_idx).value in (None, ""):
            ws.delete_cols(c_idx)


# ══════════════════════════════════════════════════════════════════════════════
# PHASE 1 단계 함수
# ══════════════════════════════════════════════════════════════════════════════

def step1_1_load_cohort(vb_cohort_path: str, site_code: str) -> dict:
    """VB 코호트 엑셀 로드 및 site_code 필터링"""
    print("\n[1-1] VB 코호트 로드 중...")
    try:
        df_cohort = pd.read_excel(vb_cohort_path, sheet_name='신장체중(예진)_전체', dtype=str)
        df_cohort = df_cohort[df_cohort['지역병원코드'] == site_code].copy()
        df_cohort['신장'] = pd.to_numeric(df_cohort['신장'], errors='coerce')
        df_cohort['체중'] = pd.to_numeric(df_cohort['체중'], errors='coerce')
        df_cohort['PatientID_key'] = df_cohort['연구등록번호'].str.strip().str.split('.').str[0]
        df_cohort = df_cohort.sort_values('내원연월', ascending=False)
        cohort_dict = {
            pid: records.to_dict('records')
            for pid, records in df_cohort.groupby('PatientID_key', sort=False)
        }
        print(f"-> {SITE} 예진 데이터 {len(cohort_dict)}명 확보.")
        return cohort_dict
    except Exception as e:
        raise IOError(f"VBcohort.xlsx 파일을 읽는 중 오류가 발생했습니다: {e}")


def step1_2_load_unique_excel(input_path: str, sheet_name: str):
    """지정 시트 로드 및 BMI·Height·Weight·SMI 컬럼 추가"""
    print(f"\n[1-2] '{sheet_name}' 시트 로드 및 구조 분석 중...")
    wb = load_workbook(input_path)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"시트 '{sheet_name}'을 찾을 수 없습니다. 존재하는 시트: {wb.sheetnames}")
    ws = wb[sheet_name]

    col_indices: dict[str, int | None] = {
        "PatientID": None, "TAMA": None, "IMATA": None,
        "BMI": None, "Height": None, "Weight": None, "SMI": None,
    }
    header_row_idx = None
    max_col: int = ws.max_column or 1

    for r_idx in range(1, 11):
        for c_idx in range(1, max_col + 1):
            val = str(ws.cell(row=r_idx, column=c_idx).value).strip()
            if val in col_indices:
                col_indices[val] = c_idx
                header_row_idx   = r_idx
        if col_indices["PatientID"] and col_indices["TAMA"]:
            break

    if not col_indices["PatientID"] or header_row_idx is None:
        raise ValueError("시트 안에서 'PatientID' 컬럼을 찾을 수 없습니다.")

    new_columns = ["Height", "Weight", "SMI"]
    if col_indices["BMI"] is None:
        new_columns.insert(0, "BMI")

    current_max_col: int = ws.max_column or max_col
    for col_name in new_columns:
        if col_indices[col_name] is None:
            current_max_col += 1
            cell = wcell(ws, header_row_idx, current_max_col)
            cell.value = col_name
            cell.font, cell.alignment = _BOLD_CENTER
            col_indices[col_name] = current_max_col
            print(f"-> '{col_name}' 컬럼 추가 ({current_max_col}번째 열)")

    return wb, ws, col_indices, header_row_idx


def step1_3_map_cohort(ws, col_indices: dict, header_row_idx: int, cohort_dict: dict):
    """PatientID 매칭 → 신장/체중/BMI/SMI 기입"""
    print("\n[1-3] VB 코호트 매칭 및 지표 기입 중...")
    pid_col    = col_indices["PatientID"]
    tama_col   = col_indices["TAMA"]
    imata_col  = col_indices["IMATA"]
    bmi_col    = col_indices["BMI"]
    height_col = col_indices["Height"]
    weight_col = col_indices["Weight"]
    smi_col    = col_indices["SMI"]
    assert pid_col is not None and tama_col is not None and imata_col is not None

    for r_idx in tqdm(range(header_row_idx + 1, ws.max_row + 1), desc="Mapping & Calculating"):
        pid_val = ws.cell(row=r_idx, column=pid_col).value
        if pid_val is None or str(pid_val).strip() == "None":
            continue
        pid_key   = str(pid_val).strip().split('.')[0]
        raw_tama  = ws.cell(row=r_idx, column=tama_col).value
        raw_imata = ws.cell(row=r_idx, column=imata_col).value
        tama  = float(raw_tama)  if raw_tama  is not None else 0.0
        imata = float(raw_imata) if raw_imata is not None else 0.0

        if pid_key not in cohort_dict:
            continue

        height, weight = None, None
        for record in cohort_dict[pid_key]:
            h = record.get('신장(cm)') or record.get('신장')
            w = record.get('체중(kg)') or record.get('체중')
            if h is None or pd.isna(h) or w is None or pd.isna(w):
                continue
            if not in_range(h, VALID_HEIGHT) or not in_range(w, VALID_WEIGHT):
                continue
            h_m = float(h) / 100.0 if float(h) > 3.0 else float(h)
            if h_m <= 0 or not in_range(float(w) / (h_m ** 2), VALID_BMI):
                continue
            height, weight = h, w
            break

        if height is not None and height_col:
            wcell(ws, r_idx, height_col).value = float(height)
        if weight is not None and weight_col:
            wcell(ws, r_idx, weight_col).value = float(weight)
        if height is not None and weight is not None and bmi_col:
            h_m = float(height) / 100.0 if float(height) > 3.0 else float(height)
            if h_m > 0:
                wcell(ws, r_idx, bmi_col).value = round(float(weight) / (h_m ** 2), 2)
        if height is not None and smi_col:
            try:
                h_m = float(height) / 100.0 if float(height) > 3.0 else float(height)
                if h_m > 0:
                    wcell(ws, r_idx, smi_col).value = round((tama - imata) / (h_m ** 2), 2)
            except ZeroDivisionError:
                pass


def step1_4_fill_bmi(ws, col_indices: dict, header_row_idx: int):
    """BMI 결측치 보완 (Height·Weight로 재계산)"""
    print("\n[1-4] BMI 결측치 보완 중...")
    bmi_col    = col_indices["BMI"]
    height_col = col_indices["Height"]
    weight_col = col_indices["Weight"]
    bmi_filled = 0
    for r_idx in range(header_row_idx + 1, ws.max_row + 1):
        if not (bmi_col and height_col and weight_col):
            break
        if ws.cell(row=r_idx, column=bmi_col).value is not None:
            continue
        h_val = ws.cell(row=r_idx, column=height_col).value
        w_val = ws.cell(row=r_idx, column=weight_col).value
        if h_val is None or w_val is None:
            continue
        try:
            h = float(h_val)
            h_m = h / 100.0 if h > 3.0 else h
            if h_m > 0:
                wcell(ws, r_idx, bmi_col).value = round(float(w_val) / (h_m ** 2), 2)
                bmi_filled += 1
        except (ValueError, ZeroDivisionError):
            pass
    print(f"-> BMI 보완 완료: {bmi_filled}명")


def step1_7_fill_smi(ws, col_indices: dict, header_row_idx: int):
    """SMI 결측치 보완 (TAMA·IMATA·Height로 재계산)"""
    print("\n[1-7] SMI 결측치 보완 중...")
    smi_col    = col_indices["SMI"]
    height_col = col_indices["Height"]
    tama_col   = col_indices["TAMA"]
    imata_col  = col_indices["IMATA"]
    smi_filled = 0
    for r_idx in range(header_row_idx + 1, ws.max_row + 1):
        if not (smi_col and height_col):
            break
        if ws.cell(row=r_idx, column=smi_col).value is not None:
            continue
        h_val  = ws.cell(row=r_idx, column=height_col).value
        t_val  = ws.cell(row=r_idx, column=tama_col).value
        im_val = ws.cell(row=r_idx, column=imata_col).value
        if h_val is None or t_val is None or im_val is None:
            continue
        try:
            h = float(h_val)
            h_m = h / 100.0 if h > 3.0 else h
            if h_m > 0:
                wcell(ws, r_idx, smi_col).value = round(
                    (float(t_val) - float(im_val)) / (h_m ** 2), 2
                )
                smi_filled += 1
        except (ValueError, ZeroDivisionError):
            pass
    print(f"-> SMI 보완 완료: {smi_filled}명")


def step1_9_save_final(wb, ws, header_row_idx: int, final_output_path: str):
    """빈 컬럼 제거 후 최종 저장"""
    print(f"\n[1-9] 최종 파일 저장 중...")
    drop_blank_columns(ws, header_row_idx)
    wb.save(final_output_path)
    wb.close()
    print(f"[PHASE 1 완료] '{final_output_path}'")


# ══════════════════════════════════════════════════════════════════════════════
# 실행
# ══════════════════════════════════════════════════════════════════════════════

def run_phase1():
    print("=" * 78)
    print(f"PHASE 1 [{SITE}]: VB 코호트 매핑 및 BMI·SMI 계산 (병원코드: {SITE_CODE})")
    print(f"입력 시트: {INPUT_SHEET}")
    print("=" * 78)

    cohort_dict = step1_1_load_cohort(VB_COHORT_PATH, SITE_CODE)
    wb, ws, col_indices, header_row_idx = step1_2_load_unique_excel(INPUT_PATH, INPUT_SHEET)
    step1_3_map_cohort(ws, col_indices, header_row_idx, cohort_dict)
    step1_4_fill_bmi(ws, col_indices, header_row_idx)
    step1_7_fill_smi(ws, col_indices, header_row_idx)
    step1_9_save_final(wb, ws, header_row_idx, FINAL_OUTPUT_PATH)

    print("\n" + "=" * 78)
    print(f"[전체 완료] {SITE} kVp_100 조영제X 데이터 처리 종료")
    print("=" * 78)


if __name__ == "__main__":
    run_phase1()
