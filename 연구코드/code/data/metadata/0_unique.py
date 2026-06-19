"""
[전체 목적]
강남 DLO(Deep Learning Output) 결과 엑셀에서 중복을 제거하고,
리포트 이미지(Osteo_Sarco_1.png)에 OCR을 적용해
CT 장비 모델명(ManufacturerModelName)·TAMA·IMATA 값을 자동 추출·보완한다.

[처리 흐름]
  STEP 1. openpyxl로 엑셀 로드 → 컬럼 위치 자동 탐색
  STEP 2. PatientID 정제 + 중복/비숫자 행 역순 제거
  STEP 3. 유니크 환자별 이중 영역 OCR (제조사 모델명 + TAMA/IMATA)
  STEP 4. 최종 저장

[용어 설명]
  DLO(Deep Learning Output) : CT 영상에서 근육 면적을 자동 산출한 결과 파일
  TAMA  : Total Abdominal Muscle Area — 복벽 근육 전체 단면적 (mm²)
  IMATA : Intramuscular Adipose Tissue Area — 근육 내 지방 단면적 (mm²)
  HYPERLINK 수식 : 엑셀 셀에 저장된 =HYPERLINK("경로", "텍스트") 형태의 수식.
                   openpyxl로 읽어야 수식 원본을 볼 수 있다.
  OCR   : Optical Character Recognition. 이미지 속 텍스트를 인식하는 기술.
  psm   : Tesseract OCR의 Page Segmentation Mode.
           --psm 3 = 전체 자동 레이아웃 분석 / --psm 6 = 균일한 표 형태 가정

[출력 컬럼 (강남_DLO_Results_Unique.xlsx)]
  1. PatientID             : 환자 고유 번호 (정수형)
  2. PatientAge            : 나이
  3. PatientSex            : 성별
  4. kVp                   : CT 관전압
  5. mAs                   : CT 관전류-시간곱
  6. ManufacturerModelName : OCR로 추출한 CT 장비 모델명 (신규 추가)
  7. SRC_Report            : 리포트 이미지 HYPERLINK 수식이 저장된 열
  8. TAMA                  : OCR로 검증/보완한 근육 단면적
  9. IMATA                 : OCR로 검증/보완한 근육 내 지방 단면적
  10. BMI                  : 체질량 지수
"""

import os
import re
import cv2
import numpy as np
import pandas as pd
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from tqdm import tqdm

# ── 환경 설정 ─────────────────────────────────────────────────────────────────

# Tesseract OCR 실행 파일 경로 (Windows 설치 기본 위치)
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

file_path   = r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\신촌\metadata\신촌_DLO_Results.xlsx"
output_path = r"C:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\신촌\metadata\신촌_DLO_Results_Unique.xlsx"

# DLO 리포트 이미지가 저장된 기반 디렉토리
DLO_BASE = r"D:\영상제공\신촌\신촌_결과"

# 대용량 처리 중 크래시로 인한 데이터 유실 방지를 위해
# 이 건수마다 엑셀에 중간 저장을 수행한다.
SAVE_INTERVAL = 50


# ── 내부 기능 함수 ────────────────────────────────────────────────────────────

def parse_dlo_img_path(formula_str, dlo_base):
    """
    엑셀 SRC_Report 셀의 HYPERLINK 수식에서 이미지 절대경로를 추출한다.

    수식 형태: =HYPERLINK("./relative/path/image.png", "링크텍스트")
    처리 순서:
      1. 정규식으로 따옴표 안의 경로 추출
      2. 백슬래시를 슬래시로 정규화
      3. 상대경로 앞의 './' 제거
      4. DLO_BASE와 결합해 절대경로 생성
    """
    if not formula_str:
        return None
    # HYPERLINK("경로", ...) 에서 첫 번째 인자(경로)를 캡처
    m = re.search(r'HYPERLINK\("([^"]+)"', str(formula_str))
    if not m:
        return None
    rel = m.group(1).replace("\\", "/").lstrip("./")       # 상대경로 정규화
    full = os.path.join(dlo_base, rel.replace("/", os.sep)) # OS 구분자로 변환
    return full


def extract_metadata_from_sarco(img_path):
    """
    Osteo_Sarco_1.png 리포트 이미지에서 세 가지 값을 OCR로 추출한다.
      - ManufacturerModelName : 전체 이미지 OCR (--psm 3)
      - TAMA, IMATA           : 하단 표 영역 크롭 OCR (--psm 6)

    [이중 영역 전략 이유]
    이미지 전체를 --psm 6(표 형태)으로 읽으면 상단 자유 텍스트 인식이 나빠지고,
    반대로 --psm 3으로 하단 표를 읽으면 열 정렬 인식이 불안정하다.
    따라서 목적에 맞게 영역을 분리해 각각 최적 모드를 적용한다.

    [이진화(Threshold) 전처리]
    OCR 전에 그레이스케일 변환 + 임계값 180 이진화를 수행한다.
    배경을 흰색(255)으로 통일해 OCR 인식률을 높이기 위함이다.

    반환: {"ManufacturerModelName": ..., "TAMA": ..., "IMATA": ...} 딕셔너리
    값이 없으면 해당 키가 빠진 빈 딕셔너리 반환.
    """
    if not img_path or not os.path.exists(img_path):
        return {}

    results = {}
    try:
        img = Image.open(img_path)
        w, h = img.size
        arr  = np.array(img)

        # 전처리: RGB → 그레이스케일 → 이진화 (배경 흰색, 텍스트 검정)
        gray   = cv2.cvtColor(arr, cv2.COLOR_RGB2GRAY)
        _, thresh = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)

        # ── OCR ①: 전체 이미지 → CT 장비 모델명 추출 ────────────────────
        # "Manufacturer Model Name : SOMATOM Force" 같은 패턴에서 모델명 부분 캡처
        full_text   = pytesseract.image_to_string(thresh, config="--psm 3")
        model_match = re.search(
            r"(?:Manufacturer\s*Model\s*Name|Model\s*Name)\s*[:|-]?\s*([^\n]+)",
            full_text, re.IGNORECASE
        )
        if model_match:
            results["ManufacturerModelName"] = model_match.group(1).strip()

        # ── OCR ②: 하단 표 영역 크롭 → TAMA/IMATA 추출 ─────────────────
        # 표는 이미지 하단 72~93% 구간에 위치한다 (경험적 좌표)
        crop      = img.crop((0, int(h * 0.72), w, int(h * 0.93)))
        crop_arr  = np.array(crop)
        crop_gray = cv2.cvtColor(crop_arr, cv2.COLOR_RGB2GRAY)
        _, crop_thresh = cv2.threshold(crop_gray, 180, 255, cv2.THRESH_BINARY)
        table_text = pytesseract.image_to_string(crop_thresh, config="--psm 6")

        # 표 각 줄에서 "TAMA  1234  2345  3456  4567" 형태 추출
        # 4개 숫자 중 마지막(인덱스 3) 값이 최종 측정값
        for line in table_text.split("\n"):
            line = line.strip()
            m = re.match(r"(TAMA|IMATA)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)", line)
            if m:
                vals = [int(m.group(i)) for i in range(2, 6)]
                results[m.group(1)] = vals[3]   # 4번째 열 = 최종 측정값

        return results
    except Exception:
        # 예외 발생 시 지금까지 수집된 결과를 그대로 반환 (부분 결과 허용)
        return results


# ── STEP 1: 엑셀 로드 및 컬럼 위치 탐색 ──────────────────────────────────────
# openpyxl을 쓰는 이유: pandas.read_excel은 HYPERLINK 수식을 읽지 못하고
# 셀 표시값(텍스트)만 반환한다. openpyxl은 수식 원본을 보존한다.

print("1. 엑셀 파일 로드 및 구조 분석 중...")
wb = load_workbook(file_path)
ws = wb.active

# 탐색할 컬럼명 → 열 인덱스 매핑 (None = 아직 미발견)
col_indices = {
    "PatientID": None,
    "SRC_Report": None,
    "TAMA": None,
    "IMATA": None,
    "ManufacturerModelName": None
}
header_row_idx = None

# 1~10행 범위에서 헤더를 찾는다 (데이터가 항상 1행에서 시작하지 않을 수 있음)
for r_idx in range(1, 11):
    for c_idx in range(1, ws.max_column + 1):
        val = str(ws.cell(row=r_idx, column=c_idx).value).strip()
        if val in col_indices:
            col_indices[val] = c_idx
            header_row_idx   = r_idx
    # 필수 컬럼 두 개가 모두 발견되면 중단
    if col_indices["PatientID"] and col_indices["SRC_Report"]:
        break

if not col_indices["PatientID"] or not col_indices["SRC_Report"]:
    raise ValueError("필수 컬럼('PatientID' 또는 'SRC_Report')을 찾을 수 없습니다.")

# ManufacturerModelName 컬럼이 없으면 시트 맨 오른쪽에 새로 추가
if col_indices["ManufacturerModelName"] is None:
    new_col_idx = ws.max_column + 1
    ws.cell(row=header_row_idx, column=new_col_idx).value = "ManufacturerModelName"
    col_indices["ManufacturerModelName"] = new_col_idx
    print(f"-> 'ManufacturerModelName' 컬럼이 없어 {new_col_idx}번째 열에 새로 추가했습니다.")

# 열 인덱스를 변수에 저장 (이후 코드에서 반복 사용)
pid_col   = col_indices["PatientID"]
src_col   = col_indices["SRC_Report"]
tama_col  = col_indices["TAMA"]
imata_col = col_indices["IMATA"]
model_col = col_indices["ManufacturerModelName"]

start_row = header_row_idx + 1   # 데이터 시작 행 (헤더 바로 다음)
max_row   = ws.max_row


# ── STEP 2: PatientID 정제 및 중복·비숫자 행 역순 제거 ───────────────────────
# [역순 제거 이유]
# openpyxl의 delete_rows()는 지운 행 아래 행의 인덱스를 즉시 당긴다.
# 아래→위 방향으로 삭제하면 이미 처리한 행의 인덱스가 바뀌지 않아 안전하다.
#
# [keep=first 구현 방법]
# seen_pids set에 PatientID를 순서대로 추가하면서,
# 이미 추가된 ID가 나오면(중복) 역순 순회이므로 → 뒤에 있는 행이 먼저 삭제됨
# → 결과적으로 가장 앞에 있는 행(첫 번째)이 남는다.

print("\n2. PatientID 정제 및 데이터 검증 구조 생성 중...")
row_info_list = []

for r_idx in range(start_row, max_row + 1):
    pid_cell = ws.cell(row=r_idx, column=pid_col)
    raw_pid  = str(pid_cell.value).strip() if pid_cell.value is not None else ""

    if not raw_pid or raw_pid == "None":
        row_info_list.append((r_idx, "None", False))
        continue

    try:
        # "12345.0" 같은 부동소수점 문자열도 정수로 변환
        pid_int       = int(float(raw_pid))
        pid_cell.value = pid_int        # 셀 값을 정수로 덮어씀 (정제)
        pid_str_key   = str(pid_int)
    except (ValueError, TypeError):
        pid_str_key = raw_pid           # 변환 실패 시 원본 문자열 유지

    # TAMA 컬럼이 숫자인지 확인 (문자열 혼입 시 해당 행 제거 대상으로 표시)
    is_tama_numeric = False
    if tama_col:
        tama_val = ws.cell(row=r_idx, column=tama_col).value
        if tama_val is not None:
            try:
                float(str(tama_val).strip())
                is_tama_numeric = True
            except ValueError:
                is_tama_numeric = False

    row_info_list.append((r_idx, pid_str_key, is_tama_numeric))

# 역순(아래→위)으로 순회하며 불필요한 행을 삭제
print("-> 역순으로 중복 행 및 TAMA 비숫자 행 일괄 제거 중...")
seen_pids          = set()
removed_by_duplicate = 0
removed_by_tama      = 0
removed_by_none      = 0

for r_idx, pid_key, is_tama_numeric in tqdm(
    reversed(row_info_list), total=len(row_info_list), desc="Cleaning Rows"
):
    if pid_key == "None":
        # PatientID가 없는 빈 행 제거
        ws.delete_rows(r_idx)
        removed_by_none += 1
        continue

    if tama_col and not is_tama_numeric:
        # TAMA에 문자열이 들어간 오염 행 제거 (예: "N/A", "error" 등)
        ws.delete_rows(r_idx)
        removed_by_tama += 1
        continue

    if pid_key in seen_pids:
        # 동일 PatientID가 이미 등록됨 → 현재 행(더 뒤에 있는 중복)을 삭제
        ws.delete_rows(r_idx)
        removed_by_duplicate += 1
    else:
        # 처음 나온 PatientID → 유지하고 seen에 등록
        seen_pids.add(pid_key)

updated_max_row    = ws.max_row
unique_data_count  = updated_max_row - start_row + 1

print("\n📊 [데이터 정제 결과 요약]")
print(f"   • 최종 유니크 행 수: {unique_data_count}개")
print("──────────────────────────────────────────────────────────────────────────────")


# ── STEP 3: 유니크 환자별 OCR 매핑 및 주기적 저장 ────────────────────────────
# 각 유니크 환자의 SRC_Report 셀에서 이미지 경로를 추출하고 OCR을 수행한다.
# 이미지 1장당 1회만 로드해 이중 OCR(전체+크롭)을 수행함으로써 I/O를 최소화한다.

print(f"\n3. OCR 분석 진행 및 실시간 매핑 중... (매 {SAVE_INTERVAL}건마다 자동 저장)")

processed_count = 0       # OCR 처리 성공 환자 수 (저장 주기 계산용)
ocr_log_limit   = 5       # 콘솔에 OCR 결과를 출력할 최대 횟수 (로그 과다 방지)
ocr_log_count   = 0

for r_idx in tqdm(range(start_row, updated_max_row + 1), desc="OCR & Data Mapping"):
    pid_cell = ws.cell(row=r_idx, column=pid_col)
    if pid_cell.value is None or str(pid_cell.value).strip() == "None":
        continue   # PatientID 없는 행은 건너뜀

    # HYPERLINK 수식에서 이미지 절대경로 추출
    src_cell = ws.cell(row=r_idx, column=src_col)
    img_path = parse_dlo_img_path(src_cell.value, DLO_BASE)

    if img_path and os.path.exists(img_path):
        ocr_results = extract_metadata_from_sarco(img_path)

        # OCR 결과를 셀에 기입 (0이 아닌 경우만 덮어씀)
        if tama_col and "TAMA" in ocr_results and ocr_results["TAMA"] != 0:
            ws.cell(row=r_idx, column=tama_col).value = int(ocr_results["TAMA"])

        if imata_col and "IMATA" in ocr_results and ocr_results["IMATA"] != 0:
            ws.cell(row=r_idx, column=imata_col).value = int(ocr_results["IMATA"])

        if model_col and "ManufacturerModelName" in ocr_results:
            ws.cell(row=r_idx, column=model_col).value = str(ocr_results["ManufacturerModelName"])

        processed_count += 1

        # 처음 5건만 콘솔에 OCR 결과 출력 (검수용)
        if ocr_log_count < ocr_log_limit:
            tqdm.write(
                f"   🔍 [OCR 확인] PatientID: {pid_cell.value} "
                f"| TAMA: {ocr_results.get('TAMA', 0)} "
                f"| IMATA: {ocr_results.get('IMATA', 0)}"
            )
            ocr_log_count += 1

        # SAVE_INTERVAL마다 중간 저장 (프로그램 중단 시 데이터 유실 방지)
        if processed_count % SAVE_INTERVAL == 0:
            wb.save(output_path)
            tqdm.write(
                f"   💾 [자동 중간 저장 완료] 현재까지 {processed_count}개 "
                f"유니크 데이터 매핑 및 저장 완료."
            )
    else:
        # 이미지 파일이 없는 경우 경고 (처음 5건만 출력)
        if ocr_log_count < ocr_log_limit and img_path:
            tqdm.write(
                f"   ⚠️ [경고] 이미지 파일 없음: {os.path.basename(img_path)} "
                f"(PatientID: {pid_cell.value})"
            )


# ── STEP 4: 최종 저장 ─────────────────────────────────────────────────────────

print(f"\n4. 최종 결과 최종 마감 저장 중...")
wb.save(output_path)   # 중간 저장과 동일한 경로에 최종 저장
wb.close()

print(f"\n[작업 완료] 주기적 자동 저장이 포함된 모든 프로세스가 성공적으로 끝났습니다.")
print(f"최종 결과 파일: '{output_path}'")
