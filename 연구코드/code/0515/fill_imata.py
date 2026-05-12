"""
신촌_merged_features.xlsx metadata-bmi_add 시트 IMATA 컬럼 자동 채우기

흐름:
1. 신촌_DLO_Results.xlsx에서 PatientID 매칭 (SRC='L3/...' 첫 번째 행)
   N열 HYPERLINK → ClariSarco Osteo_Sarco_1.png 경로 추출
2. 이미지 하단 표에서 L3 Muscle Quantity IMATA 추출
   ① RapidOCR 2x: 다자릿수 값 빠른 검출 (TAMA/NAMA/LAMA/IMATA 행)
   ② Tesseract 개별 셀: 단자릿수 값(1~9) 정밀 읽기 (PSM 7 + whitelist)
   ③ TAMA-NAMA-LAMA 계산 (각 값을 ①②로 취득 후 계산)
3. PatientID 당 L3 행 여러 개이면 첫 번째 행 사용

출력:
- 신촌_merged_features.xlsx  IMATA 컬럼 업데이트
- fill_imata_log.csv         PatientID / IMATA / method / detail
"""

import re
import os
import csv

import numpy as np
import openpyxl
import pytesseract
from PIL import Image
from rapidocr_onnxruntime import RapidOCR
from tqdm import tqdm

# ─── 경로 설정 ─────────────────────────────────────────────────────────────────
SITE = "신촌"
MERGED_FEATURES_PATH = r"c:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\\" + SITE + r"\{SITE}_merged_features.xlsx"
DLO_RESULTS_PATH     = r"D:\영상제공\\" + SITE + r"_결과\\" + SITE + r"_DLO_Results.xlsx"
DLO_BASE_DIR         = r"D:\영상제공\\" + SITE + r"\\" + SITE + r"_결과"
LOG_PATH             = r"c:\Users\jhjun\OneDrive\Desktop\2026-1_Study\연구코드\data\\" + SITE + r"\fill_imata_log.csv"

SHEET_NAME     = "metadata-bmi_add"
DLO_DATA_START = 5
COL_PID, COL_SRC, COL_REPORT = 2, 7, 13  # 0-indexed

pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
TESS_CFG = "--psm 7 -c tessedit_char_whitelist=0123456789"

ocr = RapidOCR()


# ─── 헬퍼 ──────────────────────────────────────────────────────────────────────

def parse_image_path(formula: str) -> str | None:
    if not formula:
        return None
    m = re.search(r'HYPERLINK\("([^"]+)"', str(formula))
    if not m:
        return None
    rel = m.group(1).strip("./\\")
    rel = rel.replace("/", os.sep)
    return os.path.join(DLO_BASE_DIR, rel)


def run_ocr(pil_img: Image.Image, scale: int = 2) -> list[dict]:
    """RapidOCR: scale배 확대 후 OCR, 좌표는 1x 기준으로 반환"""
    scaled = pil_img.resize((pil_img.width * scale, pil_img.height * scale), Image.LANCZOS)
    result, _ = ocr(np.array(scaled))
    if not result:
        return []
    items = []
    for bbox, text, _ in result:
        xs = [p[0] / scale for p in bbox]
        ys = [p[1] / scale for p in bbox]
        items.append({
            "x":     (min(xs) + max(xs)) / 2,
            "y_min": min(ys),
            "y_max": max(ys),
            "text":  text.strip(),
        })
    return items


def tess_read_cell(pil_img: Image.Image, x_center: float, y_min: float, y_max: float,
                   pad_x: int = 55, pad_y: int = 20, scale: int = 5) -> int | None:
    """Tesseract로 특정 셀 영역을 잘라 정수 읽기 (단자릿수에 강함)"""
    import cv2
    w, h = pil_img.size
    x1 = max(0, int(x_center - pad_x))
    x2 = min(w, int(x_center + pad_x))
    y1 = max(0, int(y_min - pad_y))
    y2 = min(h, int(y_max + pad_y))
    cell = pil_img.crop((x1, y1, x2, y2))

    # 확대 + 이진화
    big = cell.resize((cell.width * scale, cell.height * scale), Image.LANCZOS)
    gray = cv2.cvtColor(np.array(big), cv2.COLOR_RGB2GRAY)
    _, thr = cv2.threshold(gray, 180, 255, cv2.THRESH_BINARY)
    pil_thr = Image.fromarray(thr)

    text = pytesseract.image_to_string(pil_thr, config=TESS_CFG).strip()
    try:
        return int(text)
    except ValueError:
        return None


def get_label_row(items: list[dict], label: str) -> dict | None:
    """OCR 결과에서 지정 레이블 행 정보 반환"""
    matches = [it for it in items if it["text"].upper() == label.upper()]
    return matches[0] if matches else None


def find_l3_value_in_row(items: list[dict], label_row: dict,
                         l3_x: float, half_w: float,
                         tol_x: int = 60, tol_y: int = 15) -> int | None:
    """RapidOCR items에서 특정 행의 L3 열 값 추출"""
    ry_min = label_row["y_min"]
    ry_max = label_row["y_max"]
    cands = [
        it for it in items
        if it["y_min"] <= ry_max + tol_y and it["y_max"] >= ry_min - tol_y
        and it["text"].upper() != label_row["text"].upper()
        and it["x"] < half_w
        and abs(it["x"] - l3_x) < tol_x
    ]
    cands.sort(key=lambda it: abs(it["x"] - l3_x))
    for it in cands:
        try:
            return int(round(float(it["text"])))
        except ValueError:
            pass
    return None


def get_l3_value(items: list[dict], bottom: Image.Image,
                 label: str, l3_x: float, half_w: float) -> int | None:
    """
    지정 레이블 행의 L3 열 값을 RapidOCR → Tesseract 순서로 취득
    """
    row = get_label_row(items, label)
    if row is None:
        return None

    # ① RapidOCR 직접
    val = find_l3_value_in_row(items, row, l3_x, half_w)
    if val is not None:
        return val

    # ② Tesseract 개별 셀
    return tess_read_cell(bottom, l3_x, row["y_min"], row["y_max"])


def extract_l3_imata(img_path: str) -> tuple[int | None, str, str]:
    """
    ClariSarco 리포트 PNG에서 L3 Muscle Quantity IMATA 추출.
    Returns: (value, method, detail)
      method: 'direct-ocr' | 'direct-tess' | 'calc' | 'fail'
    """
    if not img_path or not os.path.exists(img_path):
        return None, "fail", f"이미지 없음: {img_path}"

    img    = Image.open(img_path)
    crop_y = int(img.height * 0.60)
    bottom = img.crop((0, crop_y, img.width, img.height))
    half_w = img.width / 2

    # ── RapidOCR (2x) ────────────────────────────────────────────────────────
    items = run_ocr(bottom, scale=2)

    # ① L3 x-위치: TAMA 행 qty 값 4번째(L3)
    l3_x: float | None = None
    tama_row = get_label_row(items, "TAMA")
    if tama_row:
        qty_vals = [
            it for it in items
            if it["y_min"] <= tama_row["y_max"] + 15
            and it["y_max"] >= tama_row["y_min"] - 15
            and it["text"].upper() != "TAMA"
            and it["x"] < half_w
        ]
        qty_vals.sort(key=lambda it: it["x"])
        if len(qty_vals) >= 4:
            l3_x = qty_vals[3]["x"]

    if l3_x is None:  # fallback: L3 헤더 텍스트
        l3_hdrs = [it for it in items if it["text"] == "L3" and it["x"] < half_w]
        if l3_hdrs:
            l3_x = l3_hdrs[0]["x"]

    if l3_x is None:
        return None, "fail", "L3 x 위치 추정 실패"

    # ② IMATA 직접 읽기 (RapidOCR)
    imata_row = get_label_row(items, "IMATA")
    if imata_row:
        val_direct = find_l3_value_in_row(items, imata_row, l3_x, half_w)
        if val_direct is not None:
            return val_direct, "direct-ocr", ""

        # ② Tesseract 개별 셀
        val_tess = tess_read_cell(bottom, l3_x, imata_row["y_min"], imata_row["y_max"])
        if val_tess is not None:
            return val_tess, "direct-tess", ""

    # ③ TAMA-NAMA-LAMA 계산 (각 값은 RapidOCR → Tesseract)
    t = get_l3_value(items, bottom, "TAMA", l3_x, half_w)
    n = get_l3_value(items, bottom, "NAMA", l3_x, half_w)
    l = get_l3_value(items, bottom, "LAMA", l3_x, half_w)

    if all(v is not None for v in (t, n, l)):
        return t - n - l, "calc", f"TAMA={t},NAMA={n},LAMA={l}"

    return None, "fail", f"TAMA={t},NAMA={n},LAMA={l}"


# ─── 메인 ──────────────────────────────────────────────────────────────────────

def main():
    print("DLO Results 로딩...")
    wb_dlo = openpyxl.load_workbook(DLO_RESULTS_PATH, read_only=True)
    ws_dlo = wb_dlo.active
    dlo_map: dict[str, str] = {}
    for row in ws_dlo.iter_rows(min_row=DLO_DATA_START, values_only=True):
        if row[COL_PID] is None:
            continue
        pid = str(row[COL_PID]).strip()
        src = str(row[COL_SRC]) if row[COL_SRC] else ""
        if src.startswith("L3") and pid not in dlo_map:
            dlo_map[pid] = row[COL_REPORT]
    wb_dlo.close()
    print(f"  L3 항목: {len(dlo_map)}명")

    print("Merged Features 로딩...")
    wb = openpyxl.load_workbook(MERGED_FEATURES_PATH)
    ws = wb[SHEET_NAME]
    hdrs     = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    pid_col   = hdrs.index("PatientID") + 1
    imata_col = hdrs.index("IMATA")     + 1

    logs: list[list] = []
    updated = skipped = failed = 0

    # --- 수정된 부분: 중간 저장 주기 설정 ---
    SAVE_INTERVAL = 50 
    batch_counter = 0

    for row_idx in tqdm(range(2, ws.max_row + 1), desc="처리 중"):
        pid_cell   = ws.cell(row=row_idx, column=pid_col)
        imata_cell = ws.cell(row=row_idx, column=imata_col)

        if pid_cell.value is None:
            break
        if imata_cell.value is not None:
            skipped += 1
            continue

        pid = str(pid_cell.value).strip()

        if pid not in dlo_map:
            logs.append([pid, "", "fail", "DLO Results에 L3 행 없음"])
            failed += 1
            continue

        img_path = parse_image_path(dlo_map[pid])
        if img_path is None:
            logs.append([pid, "", "fail", "이미지 경로 파싱 실패"])
            failed += 1
            continue

        val, method, detail = extract_l3_imata(img_path)
        if val is not None:
            imata_cell.value = val
            updated += 1
            batch_counter += 1 # 실제 업데이트된 항목 카운트
            logs.append([pid, val, method, detail])
        else:
            logs.append([pid, "", "fail", detail])
            failed += 1

        # --- 수정된 부분: 50개 주기마다 저장 ---
        if batch_counter > 0 and batch_counter % SAVE_INTERVAL == 0:
            wb.save(MERGED_FEATURES_PATH)
            # 로그도 중간에 기록하고 싶다면 여기서 파일 쓰기를 수행할 수 있습니다.
            #print(f"\n[중간 저장] {updated}번째 항목 업데이트 완료 및 저장됨...")

    # 최종 저장
    print(f"\n최종 완료: 업데이트 {updated} / 이미 채워짐 {skipped} / 실패 {failed}")
    wb.save(MERGED_FEATURES_PATH)
    print(f"최종 저장 완료: {MERGED_FEATURES_PATH}")

    # 로그 저장
    with open(LOG_PATH, "w", newline="", encoding="utf-8-sig") as f:
        csv.writer(f).writerows(
            [["PatientID", "IMATA", "method", "detail"]] + logs
        )
    print(f"로그: {LOG_PATH}")


if __name__ == "__main__":
    main()
